from flask import Flask, render_template_string, request, jsonify, Response, send_file, stream_with_context
import psycopg2
import psycopg2.extras
import yaml
import io
import re
import os
import logging
import json
import time
import threading
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from pathlib import Path
import sqlite3
from fpdf import FPDF

# Global status helper (will be overwritten by tasks import if available)
def set_status(msg, is_running=True, stats=None):
    pass

try:
    from tasks import set_status as tasks_set_status, auto_pilot_task, scrape_category_task, fast_scrape_task, direct_scrape_task, direct_gov_scrape_batch
    set_status = tasks_set_status
except ImportError:
    auto_pilot_task = scrape_category_task = fast_scrape_task = direct_gov_scrape_batch = None

class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        return super().default(obj)

app = Flask(__name__)
app.json.cls = CustomJSONEncoder

PROJ_DIR = Path(__file__).parent
LOGS_DIR = PROJ_DIR / "logs"
LOGS_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler(LOGS_DIR / "dashboard.log"), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# Log environment for Railway diagnostics
PORT = os.environ.get("PORT", "5000")
RAILWAY_SERVICE = os.environ.get("RAILWAY_SERVICE_NAME", "Unknown")
logger.info(f"BOOTSTRAP: Railway Service: {RAILWAY_SERVICE} | Port: {PORT}")

@app.route("/health")
def health_check():
    """Lightweight health check for Railway"""
    return jsonify({
        "status": "healthy",
        "dashboard": "ready",
        "database": "ready" if DB_INIT_READY else "pending",
        "timestamp": int(time.time()),
        "process_type": os.environ.get("PROCESS_TYPE", "web").lower(),
        "service": "contact-scraper-dashboard"
    }), 200


@app.route("/up")
def up():
    """Detailed health check for internal status."""
    status = {"status": "ok", "db": DB_INIT_READY}
    return jsonify(status), 200


# Redis for live status (optional)
REDIS_ACTIVE = False
try:
    import redis

    REDIS_URL = os.environ.get("REDIS_URL")
    if REDIS_URL:
        redis_client = redis.Redis.from_url(REDIS_URL, socket_timeout=2)
        redis_client.ping()
        REDIS_ACTIVE = True
    else:
        redis_client = None
except Exception:
    redis_client = None


# DB Globals
USE_SQLITE = False
DB_INIT_READY = os.environ.get("DASHBOARD_DB_BOOTSTRAPPED") == "1"

# --- Watchdog System ---
class ScraperWatchdog(threading.Thread):
    """
    Idle-detection watchdog to monitor and reset stalled scraping processes.
    """
    def __init__(self, check_interval=60):
        super().__init__(daemon=True)
        self.check_interval = check_interval
        self.logger = logging.getLogger("watchdog")

    def run(self):
        self.logger.info("Watchdog active: Monitoring for idle stalls...")
        while True:
            try:
                self.check_status()
            except Exception as e:
                self.logger.error(f"Watchdog error: {e}")
            time.sleep(self.check_interval)

    def check_status(self):
        conn = None
        try:
            conn = _connect_db()
            cur = conn.cursor()
            
            # 1. Get current status
            if USE_SQLITE:
                cur.execute("SELECT value, updated_at FROM system_status WHERE key = ?", ("scraper_status",))
            else:
                cur.execute("SELECT value, updated_at FROM system_status WHERE key = %s", ("scraper_status",))
            row = cur.fetchone()
            if not row: return

            status = json.loads(row["value"])
            updated_at = row["updated_at"]
            
            # 2. If marked as running, check last log activity
            if status.get("running"):
                # If no update in 10 mins, it's likely stalled
                if datetime.now() - updated_at > timedelta(minutes=10):
                    self.logger.warning("Detected stalled scraper process. Resetting to IDLE.")
                    
                    idle_status = {"message": "Idle (Auto-Reset)", "running": False, "time": datetime.now().strftime("%H:%M:%S")}
                    val_json = json.dumps(idle_status)
                    
                    if USE_SQLITE:
                        cur.execute("INSERT OR REPLACE INTO system_status (id, key, value, updated_at) VALUES (1, 'scraper_status', ?, ?)", 
                                   (val_json, datetime.now()))
                        cur.execute("INSERT INTO scraper_logs (level, message, source, created_at) VALUES (?, ?, ?, ?)", 
                                   ("WARNING", "Watchdog: Process stalled and was auto-reset.", "WATCHDOG", datetime.now()))
                    else:
                        cur.execute("""
                            INSERT INTO system_status (id, key, value, updated_at) 
                            VALUES (1, 'scraper_status', %s, NOW())
                            ON CONFLICT (id) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
                        """, (val_json,))
                        cur.execute("INSERT INTO scraper_logs (level, message, source, created_at) VALUES (%s, %s, %s, NOW())", 
                                   ("WARNING", "Watchdog: Process stalled and was auto-reset.", "WATCHDOG"))
                    
                    conn.commit()
                    if redis_client:
                        redis_client.set("scraper_status", val_json, ex=3600)
            
            cur.close()
            conn.close()
        except Exception as e:
            if conn: conn.close()
            self.logger.error(f"Status check failed: {e}")




DB_INIT_IN_PROGRESS = False
DB_INIT_LAST_ATTEMPT = 0.0
DB_INIT_LAST_ERROR = None
DB_INIT_RETRY_SECONDS = int(os.environ.get("DATABASE_INIT_RETRY_SECONDS", "15"))
DB_STATEMENT_TIMEOUT_MS = int(os.environ.get("DATABASE_STATEMENT_TIMEOUT_MS", "8000"))
FILTER_CACHE = {}  # Stores { 'cities': (data, timestamp), ... }
FILTER_CACHE_TTL = 300  # 5 minutes


def get_cached_filter(key, query, cur, params=None):
    """Get filter values with a 5-minute TTL to prevent heavy DB scans."""
    now = time.time()
    if key in FILTER_CACHE:
        val, ts = FILTER_CACHE[key]
        if (now - ts) < FILTER_CACHE_TTL:
            return val
    
    if params:
        cur.execute(query, params)
    else:
        cur.execute(query)
    data = [r[next(iter(r.keys()))] for r in cur.fetchall()]
    FILTER_CACHE[key] = (data, now)
    return data


def db_placeholder():
    return "?" if USE_SQLITE else "%s"


def build_contact_filters(
    search_query="",
    city="",
    category="",
    source="",
    quality="",
    exclude_schools=False,
    only_schools=False,
):
    """Build a contacts WHERE clause using the active database placeholder style."""
    ph = db_placeholder()
    like_op = "LIKE" if USE_SQLITE else "ILIKE"
    where_clauses = []
    params = []

    if search_query:
        where_clauses.append(
            f"(name {like_op} {ph} OR phone {like_op} {ph} OR email {like_op} {ph} OR category {like_op} {ph} OR source {like_op} {ph} OR city {like_op} {ph} OR address {like_op} {ph} OR arn {like_op} {ph} OR license_no {like_op} {ph} OR membership_no {like_op} {ph})"
        )
        search_pattern = f"%{search_query}%"
        params.extend([search_pattern] * 10)
    if city:
        where_clauses.append(f"city {like_op} {ph}")
        params.append(f"%{city}%")
    if category:
        where_clauses.append(f"category {like_op} {ph}")
        params.append(f"%{category}%")
    if source:
        where_clauses.append(f"source {like_op} {ph}")
        params.append(f"%{source}%")
    if quality:
        where_clauses.append(f"(quality_tier = {ph} OR quality_tier IS NULL)")
        params.append(quality)

    if exclude_schools:
        where_clauses.append(f"(LOWER(category) NOT LIKE {ph} AND LOWER(source) NOT IN ('npsc', 'bsai', 'aisa'))")
        params.append("%school%")
    if only_schools:
        where_clauses.append(f"(LOWER(category) LIKE {ph} OR LOWER(source) IN ('npsc', 'bsai', 'aisa'))")
        params.append("%school%")

    return " AND ".join(where_clauses) if where_clauses else "1=1", params


def get_db_url():
    """Build the database URL from environment variables."""
    db_url = os.environ.get("DATABASE_URL")
    if db_url:
        return db_url
    config = load_config()
    db_cfg = config.get("database", {}) if isinstance(config, dict) else {}
    host = os.environ.get("DATABASE_HOST", db_cfg.get("host", "localhost"))
    port = os.environ.get("DATABASE_PORT", db_cfg.get("port", 5432))
    name = os.environ.get("DATABASE_NAME", db_cfg.get("name", "scraper_db"))
    user = os.environ.get("DATABASE_USER", db_cfg.get("user", "postgres"))
    pw = os.environ.get("DATABASE_PASSWORD", db_cfg.get("password", ""))
    return f"postgresql://{user}:{pw}@{host}:{port}/{name}"


def _connect_db(statement_timeout_ms=None):
    """Open a database connection with a short timeout so web boot stays responsive."""
    global USE_SQLITE
    
    if USE_SQLITE or not os.environ.get("DATABASE_URL"):
        try:
            # Check if Postgres is reachable even if no URL is set (localhost)
            url = get_db_url()
            if "localhost" in url:
                 conn = psycopg2.connect(url, connect_timeout=1)
                 return conn
        except Exception:
            pass
            
        # Fallback to SQLite
        USE_SQLITE = True
        conn = sqlite3.connect(PROJ_DIR / "scraper_local.db")
        conn.row_factory = sqlite3.Row
        return conn

    url = get_db_url()
    # Railway may use postgres:// — psycopg2 needs postgresql://
    if url and url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql://", 1)
    connect_timeout = int(os.environ.get("DATABASE_CONNECT_TIMEOUT", "5"))
    connect_kwargs = {
        "cursor_factory": psycopg2.extras.RealDictCursor,
        "connect_timeout": connect_timeout,
        "application_name": "dashboard",
    }
    if statement_timeout_ms:
        connect_kwargs["options"] = f"-c statement_timeout={statement_timeout_ms}"

    conn = psycopg2.connect(url, **connect_kwargs)
    conn.autocommit = True
    return conn


def ensure_db_initialized(force=False):
    """Initialize schema lazily so the web process can boot before Postgres is ready."""
    global DB_INIT_LAST_ATTEMPT

    if DB_INIT_READY:
        return True

    now = time.monotonic()
    if (
        not force
        and DB_INIT_LAST_ERROR
        and (now - DB_INIT_LAST_ATTEMPT) < DB_INIT_RETRY_SECONDS
    ):
        raise RuntimeError(f"Database not ready yet: {DB_INIT_LAST_ERROR}")

    if not init_tables():
        raise RuntimeError(
            f"Database initialization failed: {DB_INIT_LAST_ERROR or 'unknown error'}"
        )

    return True


def get_db():
    """Get a fresh database connection after lazy schema initialization."""
    # If not already ready, try to initialize it now
    if not DB_INIT_READY:
        try:
            ensure_db_initialized()
        except Exception as e:
            logger.error(f"Failed to initialize database on request: {e}")
            raise
    return _connect_db(statement_timeout_ms=DB_STATEMENT_TIMEOUT_MS)


def set_active_task_id(task_id):
    if redis_client:
        try:
            if task_id:
                redis_client.set("scraper:current_task_id", task_id)
            else:
                redis_client.delete("scraper:current_task_id")
        except Exception:
            pass
    try:
        conn = _connect_db()
        cur = conn.cursor()
        val = task_id or ""
        if USE_SQLITE:
            cur.execute("INSERT OR REPLACE INTO system_status (id, key, value, updated_at) VALUES (2, 'current_task_id', ?, ?)", 
                       (val, datetime.now()))
        else:
            cur.execute("""
                INSERT INTO system_status (id, key, value, updated_at) 
                VALUES (2, 'current_task_id', %s, NOW())
                ON CONFLICT (id) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
            """, (val,))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        logger.warning(f"Failed to set active task id in DB: {e}")


def get_active_task_id():
    if redis_client:
        try:
            tid = redis_client.get("scraper:current_task_id")
            if tid:
                return tid.decode('utf-8')
        except Exception:
            pass
    try:
        conn = _connect_db()
        cur = conn.cursor()
        if USE_SQLITE:
            cur.execute("SELECT value FROM system_status WHERE key = ?", ("current_task_id",))
        else:
            cur.execute("SELECT value FROM system_status WHERE key = %s", ("current_task_id",))
        row = cur.fetchone()
        cur.close()
        conn.close()
        if row and row["value"]:
            return row["value"]
    except Exception as e:
        logger.warning(f"Failed to get active task id from DB: {e}")
    return None



def init_tables():
    """Create tables if they don't exist."""
    global DB_INIT_READY, DB_INIT_IN_PROGRESS, DB_INIT_LAST_ATTEMPT, DB_INIT_LAST_ERROR

    if DB_INIT_IN_PROGRESS:
        return DB_INIT_READY

    DB_INIT_IN_PROGRESS = True
    DB_INIT_LAST_ATTEMPT = time.monotonic()
    try:
        conn = _connect_db()
        cur = conn.cursor()
        # Use platform-aware types (SERIAL for Postgres, AUTOINCREMENT for SQLite)
        id_type = "INTEGER PRIMARY KEY AUTOINCREMENT" if USE_SQLITE else "SERIAL PRIMARY KEY"
        
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS contacts (
                id {id_type},
                name TEXT,
                phone VARCHAR(50),
                email TEXT,
                address TEXT,
                category TEXT,
                city TEXT,
                area TEXT,
                state TEXT,
                source TEXT,
                source_url TEXT,
                phone_clean VARCHAR(50),
                email_valid BOOLEAN,
                enriched BOOLEAN,
                arn TEXT,
                license_no TEXT,
                membership_no TEXT,
                quality_score INTEGER DEFAULT 0,
                quality_tier VARCHAR(500) DEFAULT 'low',
                blockchain_ca TEXT,
                scraped_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # New: System status table for scraper state
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS system_status (
                id INTEGER PRIMARY KEY,
                key VARCHAR(100) UNIQUE,
                value TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # New: Scraper logs table for activity feed
        log_id_type = "INTEGER PRIMARY KEY AUTOINCREMENT" if USE_SQLITE else "SERIAL PRIMARY KEY"
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS scraper_logs (
                id {log_id_type},
                level VARCHAR(20),
                message TEXT,
                source VARCHAR(100),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Individual column checks and expansion for existing tables
        required_columns = {
            "name": "TEXT",
            "category": "TEXT",
            "city": "TEXT",
            "area": "TEXT",
            "state": "TEXT",
            "source": "TEXT",
            "source_url": "TEXT",
            "email": "TEXT",
            "address": "TEXT",
            "phone_clean": "VARCHAR(50)",
            "email_valid": "BOOLEAN DEFAULT FALSE",
            "enriched": "BOOLEAN DEFAULT FALSE",
            "arn": "TEXT",
            "license_no": "TEXT",
            "membership_no": "TEXT",
            "quality_score": "INTEGER DEFAULT 0",
            "quality_tier": "TEXT DEFAULT 'low'",
            "blockchain_ca": "TEXT",
            "scraped_at": "TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
        }

        for column_name, column_type in required_columns.items():
            try:
                # SQLite doesn't support ADD COLUMN IF NOT EXISTS directly until very recently
                if USE_SQLITE:
                    cur.execute(f"PRAGMA table_info(contacts)")
                    existing = [r[1] for r in cur.fetchall()]
                    if column_name not in existing:
                        cur.execute(f"ALTER TABLE contacts ADD COLUMN {column_name} {column_type}")
                else:
                    # 1. Add column if missing
                    cur.execute(f"ALTER TABLE contacts ADD COLUMN IF NOT EXISTS {column_name} {column_type}")
                    # 2. Force expansion if it already exists (Postgres only)
                    if not USE_SQLITE:
                        cur.execute(f"ALTER TABLE contacts ALTER COLUMN {column_name} TYPE {column_type.split(' ')[0]}")
            except Exception as col_err:
                logger.debug(f"Migration detail: {column_name} check/expand skipped: {col_err}")

        # Optimization: Only run heavy cleanup if the unique index is missing
        index_exists = False
        if not USE_SQLITE:
            try:
                cur.execute("""
                    SELECT count(*) FROM pg_indexes 
                    WHERE indexname = 'idx_contacts_unique_phone'
                """)
                index_exists = cur.fetchone()['count'] > 0
            except:
                pass
        else:
            try:
                cur.execute("PRAGMA index_list('contacts')")
                indices = cur.fetchall()
                index_exists = any(idx[1] == 'idx_contacts_unique_phone' for idx in indices)
            except:
                pass

        if not index_exists:
            logger.info("Deduplication index missing. Running one-time cleanup...")
            
            # 1. Ensure phone_clean has a basic index to speed up the join
            cur.execute("CREATE INDEX IF NOT EXISTS idx_tmp_phone_clean ON contacts(phone_clean)")
            
            if USE_SQLITE:
                # SQLite doesn't support USING for DELETE, it's simpler
                cur.execute("""
                    DELETE FROM contacts WHERE id NOT IN (
                        SELECT MAX(id) FROM contacts GROUP BY phone_clean
                    ) AND phone_clean IS NOT NULL
                """)
            else:
                cur.execute("""
                    DELETE FROM contacts a
                    USING contacts b
                    WHERE a.id < b.id
                    AND a.phone_clean = b.phone_clean
                    AND a.phone_clean IS NOT NULL
                """)
                
                cur.execute("""
                    DELETE FROM contacts a
                    USING contacts b
                    WHERE a.id < b.id
                    AND a.email = b.email
                    AND a.email IS NOT NULL
                """)
            
            # 3. Drop temporary index
            cur.execute("DROP INDEX IF EXISTS idx_tmp_phone_clean")
            logger.info("Cleanup completed.")

        # Constraints for Deduplication (UPSERT support)
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_contacts_unique_phone ON contacts(phone_clean) WHERE phone_clean IS NOT NULL")
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_contacts_unique_email ON contacts(email) WHERE email IS NOT NULL")
        
        cur.execute("CREATE INDEX IF NOT EXISTS idx_contacts_source ON contacts(source)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_contacts_category ON contacts(category)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_contacts_city ON contacts(city)")

        cur.execute("CREATE INDEX IF NOT EXISTS idx_scraper_logs_created ON scraper_logs(created_at DESC)")
        cur.close()
        conn.close()
        DB_INIT_READY = True
        logger.info("Database tables ready!")
        return True
    except Exception as e:
        DB_INIT_READY = False
        DB_INIT_LAST_ERROR = str(e)
        logger.warning(f"Database init deferred: {e}")
        return False
    finally:
        DB_INIT_IN_PROGRESS = False

# Start Watchdog after DB utilities are defined
watchdog = ScraperWatchdog()
watchdog.start()


def load_config():
    try:
        with open("config.yaml", "r") as f:
            return yaml.safe_load(f)
    except Exception:
        return {}


# Note: In production, entrypoint.py handles eager bootstrap.
# We skip eager init at import time in Railway environments to allow Gunicorn to bind quickly.
if not DB_INIT_READY and RAILWAY_SERVICE != "Unknown":
    logger.info(f"BOOTSTRAP: Managed mode (Railway {RAILWAY_SERVICE}). Awaiting first request for local state sync.")
elif not DB_INIT_READY:
    logger.info("BOOTSTRAP: Local/Lazy mode (RAILWAY_SERVICE=Unknown).")


HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Maysan Labs</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;700&family=Inter:wght@400;500;600;700&family=Outfit:wght@400;600;800&display=swap" rel="stylesheet">
    <style>
        :root {
            --bg-obsidian: #050608;
            --bg-sidebar: #08090d;
            --card-glass: rgba(15, 18, 25, 0.7);
            --card-glass-hover: rgba(20, 24, 35, 0.9);
            --accent-emerald: #10b981;
            --accent-blue: #3b82f6;
            --accent-amber: #f59e0b;
            --accent-red: #ef4444;
            --text-primary: #f8fafc;
            --text-secondary: #94a3b8;
            --text-muted: #475569;
            --border-muted: rgba(255,255,255,0.06);
            --border-glow: rgba(16, 185, 129, 0.2);
            --glow-emerald: rgba(16, 185, 129, 0.4);
            --shadow-sm: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
            --shadow-lg: 0 20px 25px -5px rgba(0, 0, 0, 0.1), 0 10px 10px -5px rgba(0, 0, 0, 0.04);
        }

        /* Light Theme */
        [data-theme="light"] {
            --bg-obsidian: #f8fafc;
            --bg-sidebar: #ffffff;
            --card-glass: rgba(255, 255, 255, 0.9);
            --card-glass-hover: rgba(255, 255, 255, 1);
            --accent-emerald: #059669;
            --accent-blue: #2563eb;
            --accent-amber: #d97706;
            --accent-red: #dc2626;
            --text-primary: #1e293b;
            --text-secondary: #64748b;
            --text-muted: #94a3b8;
            --border-muted: rgba(0,0,0,0.08);
            --border-glow: rgba(16, 185, 129, 0.15);
            --shadow-sm: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
            --shadow-lg: 0 20px 25px -5px rgba(0, 0, 0, 0.1);
        }

        * { box-sizing: border-box; margin: 0; padding: 0; -webkit-font-smoothing: antialiased; }
        
        body { 
            font-family: 'Inter', sans-serif; 
            background: var(--bg-obsidian); 
            color: var(--text-primary); 
            overflow-x: hidden;
            background-image: 
                radial-gradient(circle at 0% 0%, rgba(16, 185, 129, 0.08) 0%, transparent 40%),
                radial-gradient(circle at 100% 100%, rgba(59, 130, 246, 0.08) 0%, transparent 40%);
            background-attachment: fixed;
        }

        /* Typography */
        h1, h2, h3, .brand-box p { font-family: 'Outfit', sans-serif; }
        .mono { font-family: 'JetBrains Mono', monospace; }

        .brand-box { margin-bottom: 8px; padding-left: 4px; }
        .brand-box p { font-size: 18px; font-weight: 800; color: #fff; letter-spacing: -0.5px; }
        .brand-box span { font-size: 10px; text-transform: uppercase; letter-spacing: 3px; color: var(--accent-emerald); display: block; margin-top: 2px; font-weight: 600; opacity: 0.8; }

        .nav-group { display: flex; flex-direction: column; gap: 6px; }
        .nav-label { font-size: 10px; text-transform: uppercase; color: var(--text-muted); letter-spacing: 1.5px; margin: 12px 0 8px 12px; font-weight: 700; }
        
        .nav-item { 
            padding: 12px 14px; border-radius: 12px; color: var(--text-secondary); 
            text-decoration: none; font-size: 14px; font-weight: 500; transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1);
            display: flex; align-items: center; gap: 12px;
            border: 1px solid transparent;
        }
        .nav-item:hover { background: rgba(255,255,255,0.04); color: #fff; border-color: rgba(255,255,255,0.05); }
        .nav-item.active { 
            background: rgba(16, 185, 129, 0.08); 
            color: var(--accent-emerald); 
            border: 1px solid rgba(16, 185, 129, 0.15);
            box-shadow: inset 0 0 10px rgba(16, 185, 129, 0.05);
        }
        .nav-item svg { opacity: 0.6; transition: 0.2s; }
        .nav-item:hover svg, .nav-item.active svg { opacity: 1; filter: drop-shadow(0 0 5px currentColor); }

        /* Export Buttons */
        .export-btn {
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            gap: 6px;
            padding: 14px 8px;
            border-radius: 12px;
            background: rgba(255, 255, 255, 0.03);
            border: 1px solid rgba(255, 255, 255, 0.08);
            color: var(--text-secondary);
            cursor: pointer;
            transition: all 0.25s ease;
            font-weight: 600;
            font-size: 11px;
            letter-spacing: 0.5px;
        }
        .export-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(0, 0, 0, 0.3);
        }
        .export-btn:active {
            transform: translateY(0);
        }
        .export-csv {
            border-color: rgba(16, 185, 129, 0.3);
            background: linear-gradient(135deg, rgba(16, 185, 129, 0.1), transparent);
        }
        .export-csv:hover {
            background: linear-gradient(135deg, rgba(16, 185, 129, 0.2), rgba(16, 185, 129, 0.05));
            border-color: var(--accent-emerald);
            box-shadow: 0 4px 15px rgba(16, 185, 129, 0.2);
        }
        .export-csv svg { color: var(--accent-emerald); }
        .export-csv:hover svg { filter: drop-shadow(0 0 6px var(--accent-emerald)); }

        .export-excel {
            border-color: rgba(59, 130, 246, 0.3);
            background: linear-gradient(135deg, rgba(59, 130, 246, 0.1), transparent);
        }
        .export-excel:hover {
            background: linear-gradient(135deg, rgba(59, 130, 246, 0.2), rgba(59, 130, 246, 0.05));
            border-color: var(--accent-blue);
            box-shadow: 0 4px 15px rgba(59, 130, 246, 0.2);
        }
        .export-excel svg { color: var(--accent-blue); }
        .export-excel:hover svg { filter: drop-shadow(0 0 6px var(--accent-blue)); }

        .export-pdf {
            border-color: rgba(239, 68, 68, 0.3);
            background: linear-gradient(135deg, rgba(239, 68, 68, 0.1), transparent);
        }
        .export-pdf:hover {
            background: linear-gradient(135deg, rgba(239, 68, 68, 0.2), rgba(239, 68, 68, 0.05));
            border-color: var(--accent-red);
            box-shadow: 0 4px 15px rgba(239, 68, 68, 0.2);
        }
        .export-pdf svg { color: var(--accent-red); }
        .export-pdf:hover svg { filter: drop-shadow(0 0 6px var(--accent-red)); }

        .export-json {
            border-color: rgba(245, 158, 11, 0.3);
            background: linear-gradient(135deg, rgba(245, 158, 11, 0.1), transparent);
        }
        .export-json:hover {
            background: linear-gradient(135deg, rgba(245, 158, 11, 0.2), rgba(245, 158, 11, 0.05));
            border-color: var(--accent-amber);
            box-shadow: 0 4px 15px rgba(245, 158, 11, 0.2);
        }
        .export-json svg { color: var(--accent-amber); }
        .export-json:hover svg { filter: drop-shadow(0 0 6px var(--accent-amber)); }

        /* Sidebar Footer */
        .system-footer { 
            margin-top: auto; padding: 16px; background: rgba(255,255,255,0.02); 
            border-radius: 14px; border: 1px solid var(--border-muted);
            backdrop-filter: blur(4px);
        }
        .system-footer p { font-size: 9px; color: var(--text-muted); margin-bottom: 8px; text-transform: uppercase; letter-spacing: 1px; font-weight: 600; }
        .status-online { display: flex; align-items: center; gap: 10px; font-size: 13px; font-weight: 600; color: var(--accent-emerald); }
        .status-dot { 
            width: 8px; height: 8px; background: var(--accent-emerald); border-radius: 50%; 
            box-shadow: 0 0 12px var(--accent-emerald);
            animation: statusPulse 2s infinite;
        }
        @keyframes statusPulse {
            0% { transform: scale(1); opacity: 1; }
            50% { transform: scale(1.3); opacity: 0.6; }
            100% { transform: scale(1); opacity: 1; }
        }

        /* Layout Wrapper */
        .layout-wrapper { 
            display: grid; 
            grid-template-columns: 180px 1fr; 
            min-height: 100vh; 
            width: 100%;
        }
        .sidebar { 
            background: var(--bg-sidebar); 
            border-right: 1px solid var(--border-muted); 
            padding: 24px 16px; 
            display: flex; 
            flex-direction: column; 
            gap: 24px; 
            height: 100vh; 
            position: sticky; 
            top: 0; 
            z-index: 100;
            overflow-y: auto;
        }
        .main-view { 
            padding: 40px; 
            min-width: 0; 
            position: relative;
            z-index: 1;
            display: flex;
            flex-direction: column;
            gap: 32px;
        }
        .header-row { display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px; }
        .header-row h2 { font-size: 28px; font-weight: 800; letter-spacing: -1px; background: linear-gradient(to right, #fff, var(--text-secondary)); -webkit-background-clip: text; -webkit-text-fill-color: transparent; }
        
        /* HUD Components */
        .stats-hud { display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; }
        .stat-card { 
            background: var(--card-glass); padding: 18px 20px; border-radius: 16px; 
            border: 1px solid var(--border-muted);
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            backdrop-filter: blur(12px);
            position: relative;
            overflow: hidden;
            min-height: 110px;
            display: flex;
            flex-direction: column;
        }
        .stat-card:hover { 
            background: var(--card-glass-hover); 
            border-color: var(--border-glow); 
            transform: translateY(-4px);
            box-shadow: var(--shadow-lg), 0 0 20px rgba(16, 185, 129, 0.05);
        }
        .stat-card::after {
            content: ''; position: absolute; top: 0; left: -100%; width: 100%; height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255,255,255,0.03), transparent);
            transition: 0.5s;
        }
        .stat-card:hover::after { left: 100%; }
        
        .stat-card .label { font-size: 10px; text-transform: uppercase; color: var(--text-secondary); letter-spacing: 1.2px; margin-bottom: 8px; display: block; font-weight: 700; }
        .stat-card .value { font-size: 26px; font-weight: 800; }
        .stat-card.emerald .value { color: var(--accent-emerald); text-shadow: 0 0 20px rgba(16, 185, 129, 0.3); }
        .stat-card.blue .value { color: var(--accent-blue); text-shadow: 0 0 20px rgba(59, 130, 246, 0.3); }
        .stat-card.amber .value { color: var(--accent-amber); text-shadow: 0 0 20px rgba(245, 158, 11, 0.3); }
        .stat-card .sub-text { font-size: 11px; color: var(--text-muted); margin-top: 6px; display: flex; align-items: center; gap: 4px; }
        .stat-card .sub-text svg { width: 12px; height: 12px; }

        .content-grid { display: flex; flex-direction: column; gap: 32px; }
        .glass-card { 
            background: var(--card-glass); border-radius: 24px; border: 1px solid var(--border-muted); 
            padding: 32px; backdrop-filter: blur(12px);
            box-shadow: var(--shadow-lg);
        }
        
        /* Charts */
        .charts-row { display: grid; grid-template-columns: repeat(3, 1fr); gap: 20px; }
        .chart-card { 
            background: var(--card-glass); border-radius: 20px; border: 1px solid var(--border-muted); 
            padding: 20px; min-height: 220px; backdrop-filter: blur(8px);
            transition: 0.3s;
        }
        .chart-card:hover { border-color: rgba(255,255,255,0.1); background: var(--card-glass-hover); }
        .chart-card p { font-size: 10px; text-transform: uppercase; color: var(--text-secondary); margin-bottom: 16px; letter-spacing: 1.5px; font-weight: 700; }
        .chart-container { position: relative; height: 160px; width: 100%; }
        
        /* Terminal & Feed */
        .terminal-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 16px; }
        .terminal-header h3 { font-size: 14px; text-transform: uppercase; letter-spacing: 2px; color: var(--text-secondary); }
        .terminal { 
            background: #08090c; border-radius: 16px; padding: 20px; height: 200px; overflow-y: auto;
            font-family: 'JetBrains Mono', monospace; font-size: 11px; line-height: 1.8;
            border: 1px solid var(--border-muted);
            scrollbar-width: thin;
            scrollbar-color: var(--accent-emerald) transparent;
            box-shadow: inset 0 2px 10px rgba(0,0,0,0.5);
        }
        .log-entry { margin-bottom: 6px; padding: 4px 12px; border-radius: 6px; transition: 0.2s; border-left: 2px solid transparent; }
        .log-entry:hover { background: rgba(255,255,255,0.03); border-left-color: var(--border-muted); }
        .log-time { color: var(--text-muted); font-size: 10px; min-width: 80px; }
        .log-src { color: var(--accent-blue); font-weight: 700; min-width: 90px; }
        .log-msg { color: #e2e8f0; }
        .log-msg.ERROR { color: var(--accent-red); font-weight: 600; }
        .log-msg.SUCCESS { color: var(--accent-emerald); font-weight: 600; }

        /* Controls */
        .input-group label { display: block; font-size: 10px; text-transform: uppercase; color: var(--text-secondary); margin-bottom: 6px; letter-spacing: 1.2px; font-weight: 700; }
        .input-group input, .input-group select { 
            width: 100%; background: #08090c; border: 1px solid var(--border-muted); padding: 10px 14px; 
            border-radius: 10px; color: #fff; font-size: 13px; outline: none; transition: all 0.2s;
            box-shadow: inset 0 2px 4px rgba(0,0,0,0.2);
        }
        .input-group input:focus, .input-group select:focus { border-color: var(--accent-emerald); box-shadow: 0 0 0 4px rgba(16, 185, 129, 0.1); }
        .input-group select { appearance: none; cursor: pointer; background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='24' height='24' viewBox='0 0 24 24' fill='none' stroke='%2394a3b8' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpolyline points='6 9 12 15 18 9'%3E%3C/polyline%3E%3C/svg%3E"); background-repeat: no-repeat; background-position: right 14px center; background-size: 18px; }

        .btn { 
            padding: 10px 20px; border-radius: 10px; font-weight: 700; cursor: pointer; border: none; font-size: 12px;
            text-transform: uppercase; letter-spacing: 1.2px; transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1); 
            display: inline-flex; align-items: center; gap: 8px;
            box-shadow: var(--shadow-sm);
        }
        .btn-primary { background: var(--accent-emerald); color: #064e3b; position: relative; overflow: hidden; }
        .btn-primary:hover { transform: scale(1.02); box-shadow: 0 0 25px var(--glow-emerald); background: #10c991; }
        .btn-primary::after { content: ''; position: absolute; top: -50%; left: -50%; width: 200%; height: 200%; background: radial-gradient(circle, rgba(255,255,255,0.2) 0%, transparent 70%); opacity: 0; transition: 0.5s; }
        .btn-primary:hover::after { opacity: 1; }

        .btn-outline { background: rgba(255,255,255,0.03); border: 1px solid var(--border-muted); color: var(--text-primary); }
        .btn-outline:hover { background: rgba(255,255,255,0.06); border-color: var(--text-secondary); color: #fff; }
        .btn-danger { background: transparent; border: 1px solid var(--accent-red); color: var(--accent-red); }
        .btn-danger:hover { background: var(--accent-red); color: #fff; box-shadow: 0 0 15px rgba(239, 68, 68, 0.4); }
        .btn-sm { padding: 8px 16px; font-size: 11px; }

        /* HUD Table */
        .table-wrap { 
            background: rgba(8, 9, 12, 0.5); border-radius: 18px; overflow: hidden; 
            border: 1px solid var(--border-muted); box-shadow: inset 0 0 20px rgba(0,0,0,0.2);
        }
        
        /* Controls Card */
        .controls-card {
            background: var(--card-glass);
            border-radius: 20px;
            border: 1px solid var(--border-muted);
            padding: 24px;
            backdrop-filter: blur(12px);
            display: flex;
            flex-direction: column;
            gap: 20px;
        }
        
        /* Search Bar */
        .search-bar-wrapper {
            display: flex;
            align-items: center;
            background: #08090c;
            border: 1px solid var(--border-muted);
            border-radius: 12px;
            padding: 4px 12px;
            gap: 10px;
            transition: all 0.2s;
        }
        .search-bar-wrapper:focus-within {
            border-color: var(--accent-emerald);
            box-shadow: 0 0 0 3px rgba(16, 185, 129, 0.1);
        }
        .search-bar-wrapper svg {
            color: var(--text-muted);
            flex-shrink: 0;
        }
        .search-bar-wrapper input {
            flex: 1;
            background: transparent;
            border: none;
            outline: none;
            color: #fff;
            font-size: 14px;
            padding: 10px 0;
        }
        .search-bar-wrapper input::placeholder {
            color: var(--text-muted);
        }
        .search-btn {
            background: var(--accent-emerald);
            color: #064e3b;
            border: none;
            padding: 8px 16px;
            border-radius: 10px;
            font-weight: 700;
            font-size: 12px;
            cursor: pointer;
            transition: all 0.2s;
        }
        .search-btn:hover {
            background: #10c991;
            transform: scale(1.02);
        }
        
        /* Filter Row */
        .filter-row {
            display: flex;
            flex-wrap: wrap;
            gap: 12px;
            align-items: flex-end;
        }
        .filter-row .input-group {
            flex: 1 1 130px;
            min-width: 0;
        }
        .filter-actions {
            display: flex;
            align-items: flex-end;
            gap: 10px;
            flex: 0 0 auto;
            margin-left: auto;
        }
        
        /* Quick Filters */
        .quick-filters {
            display: flex;
            align-items: center;
            gap: 8px;
            flex-wrap: wrap;
        }
        .quick-label {
            font-size: 11px;
            text-transform: uppercase;
            color: var(--text-muted);
            letter-spacing: 1px;
            font-weight: 700;
            margin-right: 4px;
        }
        .quick-btn {
            background: rgba(255,255,255,0.04);
            border: 1px solid var(--border-muted);
            color: var(--text-secondary);
            padding: 8px 14px;
            border-radius: 8px;
            font-size: 12px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.2s;
        }
        .quick-btn:hover {
            background: rgba(16, 185, 129, 0.1);
            border-color: var(--accent-emerald);
            color: var(--accent-emerald);
        }
        .clear-btn {
            background: rgba(239, 68, 68, 0.1);
            border: 1px solid rgba(239, 68, 68, 0.2);
            color: var(--accent-red);
            padding: 8px 14px;
            border-radius: 8px;
            font-size: 12px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.2s;
        }
        .clear-btn:hover {
            background: rgba(239, 68, 68, 0.2);
            border-color: var(--accent-red);
        }
        
        /* Table Section */
        .table-section {
            display: flex;
            flex-direction: column;
            gap: 16px;
        }
        .table-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .table-header h3 {
            font-size: 14px;
            text-transform: uppercase;
            letter-spacing: 2px;
            color: var(--text-secondary);
            font-weight: 700;
        }
        .table-actions {
            display: flex;
            align-items: center;
            gap: 12px;
        }
        .record-count {
            font-size: 12px;
            color: var(--text-muted);
            font-weight: 500;
        }
        
        /* Page Title */
        .page-title p {
            font-size: 13px;
            color: var(--text-muted);
            margin-top: 4px;
            font-weight: 500;
        }
        
        table { width: 100%; border-collapse: separate; border-spacing: 0; }
        th { 
            background: rgba(15, 18, 25, 0.8); padding: 12px 16px; text-align: left; 
            font-size: 10px; text-transform: uppercase; color: var(--text-muted); 
            letter-spacing: 1.2px; font-weight: 800; border-bottom: 1px solid var(--border-muted);
            position: sticky; top: 0; z-index: 10; backdrop-filter: blur(8px);
        }
        td { padding: 12px 16px; border-bottom: 1px solid var(--border-muted); font-size: 13px; color: var(--text-primary); transition: 0.2s; }
        tr:last-child td { border-bottom: none; }
        tr:hover td { background: rgba(16, 185, 129, 0.03); color: #fff; }
        tbody tr:nth-child(even) td { background: rgba(255,255,255,0.015); }
        td.cell-truncate { max-width: 160px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
        
        .badge { padding: 5px 10px; border-radius: 8px; font-size: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px; }
        .badge-src { background: rgba(59, 130, 246, 0.1); color: var(--accent-blue); border: 1px solid rgba(59, 130, 246, 0.2); }
        .total-count { color: var(--text-primary); font-weight: 800; }
        
        /* Pagination */
        .pagination {
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding-top: 24px;
            margin-top: 24px;
            border-top: 1px solid var(--border-muted);
            flex-wrap: wrap;
            gap: 16px;
        }
        .pagination-info {
            font-size: 13px;
            color: var(--text-muted);
            font-weight: 500;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        .pagination-info span { color: var(--text-secondary); font-weight: 700; }
        .pagination-btns { display: flex; gap: 6px; flex-wrap: wrap; }
        .pagination-btn {
            min-width: 40px;
            height: 40px;
            padding: 0 16px;
            border-radius: 10px;
            border: 1px solid var(--border-muted);
            background: rgba(255,255,255,0.03);
            color: var(--text-secondary);
            font-size: 13px;
            cursor: pointer;
            transition: all 0.2s ease;
            font-weight: 600;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        .pagination-btn:hover:not(:disabled) {
            border-color: var(--accent-emerald);
            color: var(--accent-emerald);
            background: rgba(16, 185, 129, 0.08);
            transform: translateY(-1px);
        }
        .pagination-btn.active {
            background: var(--accent-emerald);
            color: #000;
            border-color: var(--accent-emerald);
            box-shadow: 0 4px 12px rgba(16, 185, 129, 0.3);
        }
        .pagination-btn:disabled {
            opacity: 0.3;
            cursor: not-allowed;
            transform: none !important;
        }
        .pagination-btn.icon-btn {
            min-width: 40px;
            padding: 0;
        }

        /* Table Styling */
        .lead-row {
            transition: all 0.2s ease;
        }
        .lead-row:hover {
            background: rgba(16, 185, 129, 0.04);
        }
        .lead-row:hover td {
            color: #fff;
        }

        .score-wrapper {
            display: flex;
            align-items: center;
            gap: 10px;
        }
        .score-bar {
            width: 64px;
            height: 8px;
            background: rgba(255,255,255,0.08);
            border-radius: 3px;
            overflow: hidden;
        }
        .score-fill {
            height: 100%;
            border-radius: 3px;
            transition: width 0.3s ease;
        }
        .score-value {
            font-size: 11px;
            font-weight: 700;
            color: var(--text-secondary);
            min-width: 35px;
        }

        .action-btn {
            width: 32px;
            height: 32px;
            border-radius: 8px;
            border: 1px solid var(--border-muted);
            background: rgba(255,255,255,0.03);
            color: var(--text-muted);
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            transition: all 0.2s;
        }
        .action-btn:hover {
            background: rgba(255,255,255,0.08);
            border-color: var(--accent-emerald);
            color: var(--accent-emerald);
        }

        /* Theme Toggle Button */
        .theme-toggle {
            width: 40px;
            height: 40px;
            border-radius: 12px;
            border: 1px solid var(--border-muted);
            background: var(--card-glass);
            color: var(--text-secondary);
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            transition: all 0.2s;
            backdrop-filter: blur(10px);
        }
        .theme-toggle:hover {
            border-color: var(--accent-emerald);
            color: var(--accent-emerald);
            transform: rotate(15deg);
        }
        
        /* Value Animation for Stats */
        .stat-card .value.animate {
            animation: valuePop 0.4s ease;
        }
        @keyframes valuePop {
            0% { transform: scale(1); }
            50% { transform: scale(1.1); color: var(--accent-emerald); }
            100% { transform: scale(1); }
        }
        
        /* Row Checkbox */
        .row-checkbox {
            width: 18px;
            height: 18px;
            border-radius: 4px;
            border: 1px solid var(--border-muted);
            background: transparent;
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            transition: all 0.2s;
        }
        .row-checkbox:hover {
            border-color: var(--accent-emerald);
        }
        .row-checkbox.checked {
            background: var(--accent-emerald);
            border-color: var(--accent-emerald);
        }
        .row-checkbox.checked svg {
            display: block;
        }
        .row-checkbox svg {
            display: none;
            width: 12px;
            height: 12px;
            color: #000;
        }
        .select-all-row {
            background: rgba(16, 185, 129, 0.05) !important;
        }

        /* Loading Spinner */
        .loading-spinner {
            width: 24px;
            height: 24px;
            border: 3px solid rgba(255,255,255,0.1);
            border-top-color: var(--accent-emerald);
            border-radius: 50%;
            animation: spin 0.8s linear infinite;
            display: inline-block;
        }
        @keyframes spin {
            to { transform: rotate(360deg); }
        }
        
        /* Skeleton Loading */
        .skeleton-row td { padding: 18px 20px; }
        .skeleton-cell {
            height: 16px;
            background: linear-gradient(90deg, rgba(255,255,255,0.03) 25%, rgba(255,255,255,0.08) 50%, rgba(255,255,255,0.03) 75%);
            background-size: 200% 100%;
            animation: shimmer 1.5s infinite;
            border-radius: 4px;
        }
        .skeleton-text { width: 60%; }
        .skeleton-badge { width: 70px; height: 22px; border-radius: 8px; }
        .skeleton-score { width: 80px; }
        @keyframes shimmer {
            0% { background-position: 200% 0; }
            100% { background-position: -200% 0; }
        }

        /* Filter Chips */
        .filter-chips {
            display: flex;
            gap: 6px;
            flex-wrap: wrap;
            padding: 2px 0 0 0;
            min-height: 0;
        }
        .filter-chip {
            display: inline-flex;
            align-items: center;
            gap: 4px;
            padding: 3px 10px;
            border-radius: 20px;
            background: rgba(59, 130, 246, 0.1);
            border: 1px solid rgba(59, 130, 246, 0.2);
            color: var(--accent-blue, #60a5fa);
            font-size: 11px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.15s;
            user-select: none;
        }
        .filter-chip:hover {
            background: rgba(239, 68, 68, 0.15);
            border-color: rgba(239, 68, 68, 0.3);
            color: var(--accent-red);
        }
        .filter-chip:hover svg { stroke: var(--accent-red); }
        .filter-chip svg { width: 12px; height: 12px; flex-shrink: 0; }
        .filter-sort-hint {
            font-size: 10px;
            color: var(--text-muted);
            padding: 3px 0;
            margin-left: 4px;
            font-style: italic;
        }
        
        /* Tooltip */
        .tooltip {
            position: relative;
        }
        .tooltip::after {
            content: attr(data-tip);
            position: absolute;
            bottom: calc(100% + 8px);
            left: 50%;
            transform: translateX(-50%);
            background: var(--bg-sidebar);
            border: 1px solid var(--border-muted);
            padding: 6px 12px;
            border-radius: 8px;
            font-size: 11px;
            white-space: nowrap;
            opacity: 0;
            visibility: hidden;
            transition: all 0.2s;
            z-index: 100;
            pointer-events: none;
        }
        .tooltip:hover::after {
            opacity: 1;
            visibility: visible;
        }
        
        /* Better Table Row Hover */
        .lead-row {
            transition: all 0.2s ease;
            cursor: pointer;
        }
        .lead-row:hover {
            background: rgba(16, 185, 129, 0.04);
            transform: translateX(2px);
        }
        .lead-row:hover td {
            color: #fff;
        }
        .lead-row:hover td:first-child {
            color: var(--accent-emerald);
        }
        
        /* Keyboard Shortcut Hints */
        .kbd {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            min-width: 20px;
            height: 20px;
            padding: 0 6px;
            background: rgba(255,255,255,0.05);
            border: 1px solid rgba(255,255,255,0.1);
            border-radius: 4px;
            font-size: 10px;
            font-family: 'JetBrains Mono', monospace;
            color: var(--text-muted);
            margin-left: 8px;
        }
        
        /* Export Button Loading */
        .export-btn.loading {
            pointer-events: none;
            opacity: 0.7;
        }
        .export-btn.loading::after {
            content: '';
            position: absolute;
            top: 50%;
            left: 50%;
            width: 16px;
            height: 16px;
            margin: -8px 0 0 -8px;
            border: 2px solid transparent;
            border-top-color: currentColor;
            border-radius: 50%;
            animation: spin 0.6s linear infinite;
        }

        /* Stats / Charts Enhancement (merged) */

/* Notification Toast */
        .toast {
            position: fixed;
            bottom: 24px;
            right: 24px;
            padding: 14px 24px;
            border-radius: 12px;
            background: var(--card-glass);
            border: 1px solid var(--accent-emerald);
            color: var(--accent-emerald);
            font-weight: 600;
            font-size: 13px;
            display: none;
            align-items: center;
            gap: 12px;
            animation: slideUp 0.3s ease;
            z-index: 1000;
            backdrop-filter: blur(16px);
            box-shadow: 0 8px 32px rgba(0,0,0,0.4), 0 0 20px rgba(16, 185, 129, 0.1);
            max-width: 350px;
        }
        .toast.error {
            border-color: var(--accent-red);
            color: var(--accent-red);
            box-shadow: 0 8px 32px rgba(0,0,0,0.4), 0 0 20px rgba(239, 68, 68, 0.1);
        }
        .toast svg {
            flex-shrink: 0;
        }
        @keyframes slideUp {
            from { transform: translateY(20px); opacity: 0; }
            to { transform: translateY(0); opacity: 1; }
        }
        @keyframes slideIn {
            from { transform: translateX(20px); opacity: 0; }
            to { transform: translateX(0); opacity: 1; }
        }

        .pulse { animation: pulse 2s infinite; }
        @keyframes pulse { 0% { opacity: 1; } 50% { opacity: 0.4; } 100% { opacity: 1; } }

        /* Custom Scrollbar */
        ::-webkit-scrollbar { width: 8px; height: 8px; }
        ::-webkit-scrollbar-track { background: transparent; }
        ::-webkit-scrollbar-thumb { background: var(--text-muted); border-radius: 10px; border: 2px solid var(--bg-obsidian); }
        ::-webkit-scrollbar-thumb:hover { background: var(--text-secondary); }

        /* Responsive Grid */
        @media (max-width: 1400px) {
            .stats-hud { grid-template-columns: repeat(2, 1fr); }
        }
        @media (max-width: 1200px) {
            .stats-hud { grid-template-columns: repeat(2, 1fr); }
            .charts-row { grid-template-columns: 1fr; }
            .filter-row .input-group { flex: 1 1 160px; }
        }
        @media (max-width: 768px) {
            .stats-hud { grid-template-columns: 1fr; }
            .charts-row { grid-template-columns: 1fr; }
            .sidebar { display: none; }
            .layout-wrapper { grid-template-columns: 1fr; }
            .filter-row { grid-template-columns: 1fr; }
            .main-view { padding: 20px; }
        }
        .spinner-sm {
            width: 14px;
            height: 14px;
            border: 2px solid rgba(255,255,255,0.2);
            border-top-color: var(--accent-emerald);
            border-radius: 50%;
            animation: spin 0.8s linear infinite;
        }
        @keyframes spin { to { transform: rotate(360deg); } }
        .action-btn-green:hover {
            background: #059669 !important;
            transform: translateY(-1px);
            box-shadow: 0 4px 12px rgba(16, 185, 129, 0.2);
        }
        .action-btn-green:active {
            transform: translateY(0);
        }
        .action-btn-red:hover {
            background: #dc2626 !important;
            transform: translateY(-1px);
            box-shadow: 0 4px 12px rgba(239, 68, 68, 0.2);
        }
        .action-btn-red:active {
            transform: translateY(0);
        }
        /* Sidebar utility classes */
        .sidebar-btn {
            background: rgba(255,255,255,0.03);
            border: 1px solid var(--border-muted);
            color: var(--text-secondary);
            cursor: pointer;
            transition: all 0.2s;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        .sidebar-btn:hover {
            background: rgba(16,185,129,0.08);
            border-color: var(--accent-emerald);
            color: var(--accent-emerald);
        }
        .sidebar-btn-sm {
            font-size: 9px;
            padding: 1px 5px;
            height: 20px;
            border-radius: 4px;
        }
        .sidebar-btn-md {
            font-size: 10px;
            font-weight: 700;
            padding: 8px;
            border-radius: 8px;
            gap: 4px;
        }
        .sidebar-btn-lg {
            font-size: 11px;
            font-weight: 700;
            padding: 8px 10px;
            border-radius: 8px;
            gap: 6px;
        }
        .sidebar-export-item {
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .sidebar-section {
            display: flex;
            flex-direction: column;
            gap: 6px;
            padding: 4px;
        }
        .sidebar-scraper-btn {
            background: rgba(59,130,246,0.06);
            border: 1px solid var(--border-muted);
            color: var(--text-primary);
            cursor: pointer;
            transition: all 0.2s;
            display: flex;
            align-items: center;
            gap: 6px;
        }
        .sidebar-scraper-btn:hover {
            background: rgba(59,130,246,0.12);
            border-color: var(--accent-blue);
        }
        .sidebar-scraper-label {
            color: var(--accent-blue);
        }
        .portal-link-blue {
            border: 1px dashed var(--accent-blue);
            background: rgba(59,130,246,0.03);
            margin-bottom: 4px;
        }
        .portal-link-blue span {
            color: var(--accent-blue);
            font-weight: 700;
        }
        .portal-link-green {
            border: 1px dashed var(--accent-emerald);
            background: rgba(16,185,129,0.03);
            margin-top: 4px;
        }
        .portal-link-green span {
            color: var(--accent-emerald);
            font-weight: 700;
        }
        .export-btn-green {
            border: 1px solid var(--accent-emerald);
            background: rgba(16,185,129,0.08);
            padding: 10px 12px;
            height: auto;
            border-radius: 12px;
            color: var(--text-secondary);
            cursor: pointer;
            display: flex;
            align-items: center;
            gap: 8px;
            transition: all 0.2s;
        }
        .export-btn-green svg { stroke: var(--accent-emerald); }
        .export-btn-green span { color: var(--accent-emerald); font-weight: 700; font-size: 13px; }
        .export-btn-green:hover { background: rgba(16,185,129,0.15); border-color: var(--accent-emerald); }
        .export-btn-red {
            border: 1px solid var(--accent-red);
            background: rgba(239,68,68,0.08);
            padding: 10px 12px;
            height: auto;
            border-radius: 12px;
            color: var(--accent-red);
            cursor: pointer;
            display: flex;
            align-items: center;
            gap: 8px;
            transition: all 0.2s;
        }
        .export-btn-red span { font-weight: 700; font-size: 13px; }
        .export-btn-red:hover { background: rgba(239,68,68,0.15); }
        .cat-label { font-size: 10px; font-weight: 700; color: var(--text-secondary); }
        .btn-group { display: flex; gap: 3px; }
        .export-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 6px; padding: 4px; }
        .scraper-nav-group { border-radius: 12px; padding: 8px; }
        .scraper-nav-label { font-size: 12px; letter-spacing: 1px; display: flex; align-items: center; gap: 4px; }
        .scraper-control-group { display: flex; flex-direction: column; gap: 8px; padding: 6px 2px; }
        .zone-label { font-size: 10px; color: var(--text-muted); font-weight: 600; display: block; margin-bottom: 3px; }
        .zone-input { width: 100%; padding: 6px 8px; border-radius: 8px; border: 1px solid var(--border-muted); background: var(--bg-secondary); color: var(--text-primary); font-size: 12px; box-sizing: border-box; }
        .zone-btn-row { display: flex; flex-wrap: wrap; gap: 4px; }
        .zone-btn { font-size: 10px; padding: 3px 8px; border-radius: 6px; border: 1px solid var(--border-muted); background: transparent; color: var(--text-muted); cursor: pointer; transition: all 0.2s; }
        .zone-btn:hover { background: rgba(16,185,129,0.1); border-color: var(--accent-emerald); color: var(--accent-emerald); }
        .zone-btn-all { font-size: 10px; padding: 3px 8px; border-radius: 6px; border: 1px solid var(--accent-emerald); background: transparent; color: var(--accent-emerald); cursor: pointer; transition: all 0.2s; }
        .zone-btn-all:hover { background: rgba(16,185,129,0.1); }
        .scrape-btn { background: var(--accent-emerald); color: #fff; border: none; padding: 10px 12px; border-radius: 8px; font-size: 12px; font-weight: 800; cursor: pointer; display: flex; align-items: center; justify-content: center; gap: 6px; letter-spacing: 0.5px; transition: all 0.2s; }
        .scrape-btn:hover { background: #059669; transform: translateY(-1px); }
        .status-text { font-size: 10px; color: var(--text-muted); text-align: center; min-height: 16px; padding: 2px 0; }
        .gov-batch-btn { background: rgba(245,158,11,0.08); border: 1px solid var(--accent-amber); color: var(--accent-amber); padding: 10px 10px; border-radius: 8px; font-size: 11px; font-weight: 800; cursor: pointer; display: flex; align-items: center; justify-content: center; gap: 6px; letter-spacing: 0.3px; transition: all 0.2s; }
        .gov-batch-btn:hover { background: rgba(245,158,11,0.15); transform: translateY(-1px); }
        .divider-line { border-top: 1px solid var(--border-muted); margin: 4px 0 2px; }

        /* Responsive */
        @media (max-width: 1024px) {
            .stats-hud { grid-template-columns: repeat(2, 1fr); }
            .charts-row { grid-template-columns: 1fr; }
            .filter-row { flex-direction: column; }
            .filter-row .input-group { flex: 1 1 100%; }
            .filter-actions { margin-left: 0; width: 100%; }
            .filter-actions .btn { flex: 1; }
            .header-row h2 { font-size: 20px; }
        }
        @media (max-width: 640px) {
            .stats-hud { grid-template-columns: 1fr; }
            .main-view { padding: 12px; gap: 20px; }
            .glass-card { padding: 16px; }
            .controls-card { padding: 16px; }
            .header-row { flex-direction: column; align-items: flex-start; gap: 8px; }
            .header-row h2 { font-size: 18px; }
        }
    </style>
</head>
<body>
    <div id="notif" class="toast">
        <svg id="notif-icon-success" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polyline points="20 6 9 17 4 12"></polyline></svg>
        <svg id="notif-icon-error" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" style="display:none;"><circle cx="12" cy="12" r="10"></circle><line x1="15" y1="9" x2="9" y2="15"></line><line x1="9" y1="9" x2="15" y2="15"></line></svg>
        <span id="notif-text"></span>
    </div>

    <div class="layout-wrapper">
        <aside class="sidebar">
            <div class="brand-box">
                <p>Maysan Labs</p>
                <span>{% if is_school_dashboard %}Schools{% else %}Financial{% endif %}</span>
            </div>
            
            <nav class="nav-group">
                <p class="nav-label">Views</p>
                {% if is_school_dashboard %}
                <a href="/" class="nav-item" style="border:1px dashed var(--accent-blue);">
                    <svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="var(--accent-blue)" stroke-width="2"><path d="M3 9l9-7 9 7v11a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2z"></path><polyline points="9 22 9 12 15 12 15 22"></polyline></svg>
                    <span style="color:var(--accent-blue);">Financial</span>
                </a>
                <a href="/schools" class="nav-item active">
                    <svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M22 10v6M2 10v6M12 2L2 10h20L12 2zM4 10v6h16v-6"></path></svg>
                    Schools
                </a>
                {% else %}
                <a href="/" class="nav-item active">
                    <svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M3 9l9-7 9 7v11a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2z"></path><polyline points="9 22 9 12 15 12 15 22"></polyline></svg>
                    Financial
                </a>
                <a href="/schools" class="nav-item" style="border:1px dashed var(--accent-emerald);">
                    <svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="var(--accent-emerald)" stroke-width="2"><path d="M22 10v6M2 10v6M12 2L2 10h20L12 2zM4 10v6h16v-6"></path></svg>
                    <span style="color:var(--accent-emerald);">Schools</span>
                </a>
                {% endif %}
            </nav>

            <nav class="nav-group">
                <p class="nav-label">Export Data</p>
                <div class="export-grid">
                    <button onclick="exportAllData('csv', this)" class="sidebar-btn sidebar-btn-md">
                        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"></path><polyline points="7 10 12 15 17 10"></polyline><line x1="12" y1="15" x2="12" y2="3"></line></svg>
                        CSV
                    </button>
                    <button onclick="exportAllData('xlsx', this)" class="sidebar-btn sidebar-btn-md">
                        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"></path><polyline points="7 10 12 15 17 10"></polyline><line x1="12" y1="15" x2="12" y2="3"></line></svg>
                        Excel
                    </button>
                    <button onclick="exportAllData('pdf', this)" class="sidebar-btn sidebar-btn-md" style="grid-column: span 2;">
                        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"></path><polyline points="14 2 14 8 20 8"></polyline><line x1="16" y1="13" x2="8" y2="13"></line><line x1="16" y1="17" x2="8" y2="17"></line></svg>
                        PDF Report
                    </button>
                </div>
            </nav>

            <div class="system-footer">
                <p>System Status</p>
                <div style="display:flex; justify-content:space-between; align-items:center;">
                    <div class="status-online">
                        <div class="status-dot"></div>
                        <span id="status-badge">ONLINE</span>
                    </div>
                    <span id="last-update-sidebar" style="font-size:10px; color:rgba(255,255,255,0.2);">--:--:--</span>
                </div>
            </div>
        </aside>

    <main class="main-view">
        <div class="header-row">
            <div class="page-title">
                {% if is_school_dashboard %}
                <h2>Schools</h2>
                <p>School lead records</p>
                {% else %}
                <h2>Dashboard{% if selected_category %}<span style="color: var(--accent-blue);"> / {{ selected_category }}s</span>{% endif %}</h2>
                <p>{% if selected_category %}{{ selected_category }} records{% else %}Financial lead records{% endif %}</p>
                {% endif %}
            </div>
            <div style="display:flex; align-items:center; gap:10px;">
                <button class="theme-toggle" onclick="toggleTheme()" title="Toggle Theme">
                    <svg id="theme-icon-sun" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="12" cy="12" r="5"></circle><line x1="12" y1="1" x2="12" y2="3"></line><line x1="12" y1="21" x2="12" y2="23"></line><line x1="4.22" y1="4.22" x2="5.64" y2="5.64"></line><line x1="18.36" y1="18.36" x2="19.78" y2="19.78"></line><line x1="1" y1="12" x2="3" y2="12"></line><line x1="21" y1="12" x2="23" y2="12"></line><line x1="4.22" y1="19.78" x2="5.64" y2="18.36"></line><line x1="18.36" y1="5.64" x2="19.78" y2="4.22"></line></svg>
                    <svg id="theme-icon-moon" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" style="display:none;"><path d="M21 12.79A9 9 0 1 1 11.21 3 7 7 0 0 0 21 12.79z"></path></svg>
                </button>
                <div class="scraping-control-group" style="display:flex; align-items:center; gap:4px; background:var(--card-glass); border:1px solid var(--border-muted); padding:3px 6px; border-radius:10px; backdrop-filter:blur(10px);">
                    <button class="action-btn-green" onclick="startScrapeGlobal()" id="global-start-btn" style="background:var(--accent-emerald); color:#fff; border:none; padding:6px 10px; border-radius:7px; font-size:10px; font-weight:800; cursor:pointer; display:flex; align-items:center; gap:4px; letter-spacing: 0.5px; transition: all 0.2s ease;">
                        <svg width="10" height="10" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="3"><polygon points="5 3 19 12 5 21 5 3"></polygon></svg>
                        START
                    </button>
                    <button class="action-btn-red" onclick="stopScrapeGlobal()" id="global-stop-btn" style="background:var(--accent-red); color:#fff; border:none; padding:6px 10px; border-radius:7px; font-size:10px; font-weight:800; cursor:pointer; display:flex; align-items:center; gap:4px; letter-spacing: 0.5px; transition: all 0.2s ease;">
                        <svg width="10" height="10" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="3"><rect x="4" y="4" width="16" height="16" rx="2"></rect></svg>
                        STOP
                    </button>
                </div>
                <div class="status-pill" style="background:var(--card-glass); border:1px solid var(--border-muted); padding:5px 10px; border-radius:10px; font-size:11px; display:flex; align-items:center; gap:5px; backdrop-filter:blur(10px);">
                    <span style="color:var(--accent-emerald); font-weight:700;">●</span>
                    <span id="live-status" style="font-weight:800; color:var(--text-secondary); letter-spacing:0.5px;">IDLE</span>
                </div>
                <div class="status-pill" style="background:var(--card-glass); border:1px solid var(--border-muted); padding:5px 10px; border-radius:10px; font-size:11px; backdrop-filter:blur(10px); display:flex; align-items:center; gap:5px;">
                    <svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="var(--text-muted)" stroke-width="2"><circle cx="12" cy="12" r="10"></circle><polyline points="12 6 12 12 16 14"></polyline></svg>
                    <span id="last-update" class="mono" style="color:var(--text-primary);">--:--:--</span>
                </div>
            </div>
        </div>

        <div id="prog-wrap" style="display:none; margin-top: -16px;">
            <div class="progress-bar-container">
                <div id="prog-bar" class="progress-bar" style="width: 0%;"></div>
            </div>
            <p style="font-size:10px; color:var(--accent-emerald); margin-top:8px; font-weight:700; letter-spacing:1px; text-align:right;">EXTRACTION IN PROGRESS...</p>
        </div>

        <div class="stats-hud">
            <div class="stat-card tooltip" data-tip="Total scraped records">
                <span class="label">Total Leads</span>
                <span class="value mono" id="stat-total">{{s.total}}</span>
                <span class="sub-text">
                    <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"></path><circle cx="9" cy="7" r="4"></circle></svg>
                    {{ s.filtered_total }} filtered
                </span>
            </div>
            <div class="stat-card emerald tooltip" data-tip="Contacts with valid phone numbers">
                <span class="label">Verified Contacts</span>
                <span class="value mono" id="stat-phone">{{s.phone}}</span>
                <span class="sub-text">
                    <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M22 16.92v3a2 2 0 0 1-2.18 2 19.79 19.79 0 0 1-8.63-3.07 19.5 19.5 0 0 1-6-6 19.79 19.79 0 0 1-3.07-8.67A2 2 0 0 1 4.11 2h3a2 2 0 0 1 2 1.72 12.84 12.84 0 0 0 .7 2.81 2 2 0 0 1-.45 2.11L8.09 9.91a16 16 0 0 0 6 6l1.27-1.27a2 2 0 0 1 2.11-.45 12.84 12.84 0 0 0 2.81.7A2 2 0 0 1 22 16.92z"></path></svg>
                    {{ s.with_phone_pct }}% coverage
                </span>
            </div>
            <div class="stat-card blue tooltip" data-tip="Contacts with valid email addresses">
                <span class="label">Digital Identity</span>
                <span class="value mono" id="stat-email">{{s.email}}</span>
                <span class="sub-text">
                    <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M4 4h16c1.1 0 2 .9 2 2v12c0 1.1-.9 2-2 2H4c-1.1 0-2-.9-2-2V6c0-1.1.9-2 2-2z"></path><polyline points="22,6 12,13 2,6"></polyline></svg>
                    {{ s.with_email_pct }}% coverage
                </span>
            </div>
            <div class="stat-card amber tooltip" data-tip="Average quality score across all records">
                <span class="label">Data Quality</span>
                <span class="value mono">{{s.avg_quality}}%</span>
                <span class="sub-text">
                    <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polygon points="12 2 15.09 8.26 22 9.27 17 14.14 18.18 21.02 12 17.77 5.82 21.02 7 14.14 2 9.27 8.91 8.26 12 2"></polygon></svg>
                    Avg score
                </span>
            </div>
        </div>

<div class="controls-card">
            <div class="search-bar-wrapper">
                <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="11" cy="11" r="8"></circle><line x1="21" y1="21" x2="16.65" y2="16.65"></line></svg>
                <input type="text" id="t-cat" placeholder="Search leads, categories, sources..." list="cats-list" value="{{selected_category or search_query}}">
                <button class="search-btn" onclick="applyFilters()">Search<span class="kbd">Enter</span></button>
            </div>

            <div class="filter-row">
                <div class="input-group">
                    <label>Location</label>
                    <input type="text" id="t-city" placeholder="City..." list="cities-list" value="{{selected_city}}">
                </div>
                <div class="input-group">
                    <label>Source</label>
                    <select id="t-source">
                        <option value="">All Sources</option>
                        {% set source_options = sources if sources else ['AMFI', 'IRDAI', 'ICAI', 'SEBI', 'BAR_COUNCIL', 'SITEMAP', 'YELLOWPAGES', 'JUSTDIAL', 'GMB'] %}
                        {% for src in source_options %}
                        <option value="{{src}}" {% if selected_source == src %}selected{% endif %}>{{ src|replace('_', ' ')|title }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div class="input-group">
                    <label>Sort By</label>
                    <select id="t-sort" onchange="applyFilters()">
                        <option value="date" {% if sort_by == 'date' %}selected{% endif %}>Newest First</option>
                        <option value="date_asc" {% if sort_by == 'date_asc' %}selected{% endif %}>Oldest First</option>
                        <option value="name" {% if sort_by == 'name' %}selected{% endif %}>Name A-Z</option>
                        <option value="name_desc" {% if sort_by == 'name_desc' %}selected{% endif %}>Name Z-A</option>
                        <option value="score" {% if sort_by == 'score' %}selected{% endif %}>Quality High-Low</option>
                        <option value="score_asc" {% if sort_by == 'score_asc' %}selected{% endif %}>Quality Low-High</option>
                        <option value="city" {% if sort_by == 'city' %}selected{% endif %}>City A-Z</option>
                        <option value="city_desc" {% if sort_by == 'city_desc' %}selected{% endif %}>City Z-A</option>
                        <option value="category" {% if sort_by == 'category' %}selected{% endif %}>Category A-Z</option>
                        <option value="source" {% if sort_by == 'source' %}selected{% endif %}>Source A-Z</option>
                    </select>
                </div>
                <div class="input-group">
                    <label>Quality</label>
                    <select id="t-quality" onchange="applyFilters()">
                        <option value="">All Qualities</option>
                        <option value="high" {% if selected_quality == 'high' %}selected{% endif %}>High (70%+)</option>
                        <option value="medium" {% if selected_quality == 'medium' %}selected{% endif %}>Medium (40-70%)</option>
                        <option value="low" {% if selected_quality == 'low' %}selected{% endif %}>Low (&lt;40%)</option>
                    </select>
                </div>
                <div class="input-group">
                    <label>Per Page</label>
                    <select id="t-limit" onchange="applyFilters()">
                        <option value="25" {% if limit == 25 %}selected{% endif %}>25</option>
                        <option value="50" {% if limit == 50 %}selected{% endif %}>50</option>
                        <option value="100" {% if limit == 100 %}selected{% endif %}>100</option>
                        <option value="250" {% if limit == 250 %}selected{% endif %}>250</option>
                    </select>
                </div>
                <div class="filter-actions">
                    <button class="btn btn-primary" id="start-btn" onclick="startCollection()">
                        <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polygon points="5 3 19 12 5 21 5 3"></polygon></svg>
                        Start Collection
                    </button>
                    <button class="btn btn-outline" id="auto-btn" onclick="startCollection(true)" title="Run continuous autonomous scrape">
                        AutoPilot
                    </button>
                    <button class="btn btn-danger" id="stop-btn" onclick="stopScraping()" style="display:none; border-color:var(--accent-red); color:var(--accent-red);">
                        Stop
                    </button>
                </div>
            </div>

            <div class="quick-filters">
                <span class="quick-label">Quick:</span>
                {% if is_school_dashboard %}
                <button class="quick-btn" onclick="setFilter('Delhi', 'Schools')">Schools Delhi</button>
                <button class="quick-btn" onclick="setFilter('Mumbai', 'Schools')">Schools Mumbai</button>
                <button class="quick-btn" onclick="setFilter('Bangalore', 'Schools')">Schools Bangalore</button>
                {% else %}
                <button class="quick-btn" onclick="setFilter('Delhi', 'Lawyers')">Lawyers Delhi</button>
                <button class="quick-btn" onclick="setFilter('Mumbai', 'CAs')">CAs Mumbai</button>
                <button class="quick-btn" onclick="setFilter('Bangalore', 'Doctors')">Doctors Bangalore</button>
                {% endif %}
                <button class="clear-btn" onclick="clearFilters()">Clear All</button>
            </div>

            <div id="active-filter-chips" class="filter-chips"></div>
        </div>

        <div class="charts-row">
            <div class="chart-card">
                <p>Source Distribution</p>
                <div class="chart-container"><canvas id="sourceChart"></canvas></div>
            </div>
            <div class="chart-card">
                <p>Lead Health & Completeness</p>
                <div style="margin-top: 10px;">
                    <div style="margin-bottom: 20px;">
                        <div style="display:flex; justify-content:space-between; margin-bottom:8px; font-size:11px; font-weight:700; color:var(--text-secondary);">
                            <span>PHONE VERIFICATION</span>
                            <span class="mono">{{s.with_phone_pct}}%</span>
                        </div>
                        <div class="progress-bar-container"><div class="progress-bar" style="width:{{s.with_phone_pct}}%; background:var(--accent-emerald);"></div></div>
                    </div>
                    <div style="margin-bottom: 20px;">
                        <div style="display:flex; justify-content:space-between; margin-bottom:8px; font-size:11px; font-weight:700; color:var(--text-secondary);">
                            <span>DIGITAL REACH (EMAIL)</span>
                            <span class="mono">{{s.with_email_pct}}%</span>
                        </div>
                        <div class="progress-bar-container"><div class="progress-bar" style="width:{{s.with_email_pct}}%; background:var(--accent-blue);"></div></div>
                    </div>
                    <div>
                        <div style="display:flex; justify-content:space-between; margin-bottom:8px; font-size:11px; font-weight:700; color:var(--text-secondary);">
                            <span>AVG DATA FIDELITY</span>
                            <span class="mono">{{s.avg_quality}}%</span>
                        </div>
                        <div class="progress-bar-container"><div class="progress-bar" style="width:{{s.avg_quality}}%; background:var(--accent-amber);"></div></div>
                    </div>
                </div>
            </div>
            <div class="chart-card">
                <p>Top Categories</p>
                <div class="chart-container"><canvas id="categoryChart"></canvas></div>
            </div>
        </div>

        <div class="chart-card" style="margin-top: -12px;">
            <p>Growth Trend (Last 7 Days)</p>
            <div class="chart-container" style="height: 120px;"><canvas id="growthChart"></canvas></div>
        </div>

        <div class="content-grid">
            <div class="glass-card">
                <div class="table-section">
                    <div class="table-header">
                        <h3>Lead Records</h3>
                        <div class="table-actions">
                            <span class="record-count">{{ contacts|length }} of {{ s.total }} records</span>
                        </div>
                    </div>
                    <div class="table-wrap">
                        <table>
                            <thead>
                                <tr>
                                    <th style="width:40px;">
                                        <div class="row-checkbox" onclick="toggleSelectAll(this)" title="Select All">
                                            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="3"><polyline points="20 6 9 17 4 12"></polyline></svg>
                                        </div>
                                    </th>
                                    <th style="width:40px;">#</th>
                                    <th>Lead Name</th>
                                    <th>Phone</th>
                                    <th>Email</th>
                                    <th>Category</th>
                                    <th>City</th>
                                    <th>Source</th>
                                    <th style="width:100px;">Score</th>
                                    <th style="width:60px;">Actions</th>
                                </tr>
                            </thead>
                            <tbody id="leads-tbody">
                                {% for c in contacts %}
                                <tr class="lead-row">
                                    <td>
                                        <div class="row-checkbox" onclick="toggleRow(this)">
                                            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="3"><polyline points="20 6 9 17 4 12"></polyline></svg>
                                        </div>
                                    </td>
                                    <td style="color:var(--text-muted); font-size:11px;">{{ loop.index + (page - 1) * limit }}</td>
                                    <td class="cell-truncate" style="font-weight:700; font-family:'Outfit',sans-serif; color:#fff;">{{c.name}}</td>
                                    <td class="mono cell-truncate" style="font-size:12px;">{{c.phone or '---'}}</td>
                                    <td class="mono cell-truncate" style="color:var(--accent-blue); font-size:11px;">{{c.email or '---'}}</td>
                                    <td class="cell-truncate" style="font-size:12px; font-weight:500;">{{c.category}}</td>
                                    <td style="font-size:12px; color:var(--text-secondary);">{{c.city or '---'}}</td>
                                    <td><span class="badge badge-src">{{c.source}}</span></td>
                                    <td>
                                        <div class="score-wrapper">
                                            <div class="score-bar"><div class="score-fill" style="width:{{c.quality_score}}%; background:{{ 'var(--accent-emerald)' if c.quality_score > 70 else 'var(--accent-blue)' if c.quality_score > 40 else 'var(--accent-red)' }};"></div></div>
                                            <span class="mono score-value">{{c.quality_score}}%</span>
                                        </div>
                                    </td>
                                    <td>
                                        <div style="display:flex; gap:4px;">
                                            <button class="action-btn" title="Copy" data-copy="{{ (c.phone or c.email or '')|e }}" onclick="copyLead(this.dataset.copy)">
                                                <svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="9" y="9" width="13" height="13" rx="2"></rect><path d="M5 15H4a2 2 0 0 1-2-2V4a2 2 0 0 1 2-2h9a2 2 0 0 1 2 2v1"></path></svg>
                                            </button>
                                        </div>
                                    </td>
                                </tr>
                                {% endfor %}
                            </tbody>
                        </table>
                    </div>
                    
                    <div class="pagination" id="pagination-wrapper">
                        <div class="pagination-info">
                            <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"></path><circle cx="9" cy="7" r="4"></circle><path d="M23 21v-2a4 4 0 0 0-3-3.87"></path><path d="M16 3.13a4 4 0 0 1 0 7.75"></path></svg>
                            <span><span>{% set start_range = (page - 1) * limit + 1 %}{{ start_range }}-{{ start_range + contacts|length - 1 }}</span> of <span class="total-count">{{ "{:,}".format(s.filtered_total) }}</span> leads</span>
                            <span style="color:var(--border-muted);">|</span>
                            <span>Page <span>{{ page }}</span> of <span>{{ total_pages }}</span></span>
                        </div>
                        <div class="pagination-btns" id="pagination-btns-container">
                            <button class="pagination-btn icon-btn" onclick="goToPage(1)" {% if page <= 1 %}disabled{% endif %} title="First">
                                <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polyline points="11 17 6 12 11 7"></polyline><polyline points="18 17 13 12 18 7"></polyline></svg>
                            </button>
                            <button class="pagination-btn" onclick="changePage(-1)" {% if page <= 1 %}disabled{% endif %}>Prev</button>
                            
                            {% set start_p = [1, page - 2]|max %}
                            {% set end_p = [total_pages, start_p + 4]|min %}
                            {% set start_p = [1, end_p - 4]|max %}
                            
                            {% for p in range(start_p, end_p + 1) %}
                            <button class="pagination-btn {% if p == page %}active{% endif %}" onclick="goToPage({{ p }})">{{ p }}</button>
                            {% endfor %}

                            <button class="pagination-btn" onclick="changePage(1)" {% if page >= total_pages %}disabled{% endif %}>Next</button>
                            <button class="pagination-btn icon-btn" onclick="goToPage({{ total_pages }})" {% if page >= total_pages %}disabled{% endif %} title="Last">
                                <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polyline points="13 17 18 12 13 7"></polyline><polyline points="6 17 11 12 6 7"></polyline></svg>
                            </button>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </main>
</div>

    <datalist id="cities-list">{% for c in cities %}<option value="{{c}}">{% endfor %}</datalist>
    <datalist id="cats-list">{% for c in categories %}<option value="{{c}}">{% endfor %}</datalist>    <script>
        window.isSchoolDashboard = {% if is_school_dashboard %}true{% else %}false{% endif %};
        // CORE NAVIGATION FUNCTIONS (Defined early)
        window.showNotif = function(msg, dur, isError) {
            if (dur === undefined) dur = 3000;
            const n = document.getElementById('notif');
            const t = document.getElementById('notif-text');
            const errIcon = document.getElementById('notif-icon-error');
            const sucIcon = document.getElementById('notif-icon-success');
            if (n && t) {
                t.innerText = msg;
                n.style.display = 'flex';
                if (isError) {
                    n.classList.add('error');
                    if (errIcon) errIcon.style.display = 'block';
                    if (sucIcon) sucIcon.style.display = 'none';
                } else {
                    n.classList.remove('error');
                    if (errIcon) errIcon.style.display = 'none';
                    if (sucIcon) sucIcon.style.display = 'block';
                }
                setTimeout(function() { n.style.display = 'none'; }, dur);
            }
        };

        window.currentPage = parseInt("{{page}}") || 1;
        window.totalPages = parseInt("{{total_pages}}") || 1;
        window.pageSize = parseInt("{{limit}}") || 50;

        window.changePage = function(delta) {
            window.goToPage(window.currentPage + delta);
        };
        
        window.goToPage = function(p) {
            if (p < 1 || p > window.totalPages) return;
            
            const city = document.getElementById('t-city')?.value || "";
            const q = document.getElementById('t-cat')?.value || "";
            const source = document.getElementById('t-source')?.value || "";
            const sort = document.getElementById('t-sort')?.value || 'date';
            const quality = document.getElementById('t-quality')?.value || "";
            const limit = document.getElementById('t-limit')?.value || window.pageSize || '50';
            
            const url = new URL(window.location.origin + window.location.pathname);
            if (city) url.searchParams.set('city', city);
            if (q) url.searchParams.set('q', q);
            if (source) url.searchParams.set('source', source);
            if (sort) url.searchParams.set('sort', sort);
            if (quality) url.searchParams.set('quality', quality);
            
            // Preserve category parameter if present in current URL
            const currentUrl = new URL(window.location.href);
            const category = currentUrl.searchParams.get('category');
            if (category) url.searchParams.set('category', category);
            
            url.searchParams.set('page', p);
            url.searchParams.set('limit', limit);
            window.loadLeads(url.toString(), true);
        };

        window.applyFilters = function() {
            const city = document.getElementById('t-city').value;
            const q = document.getElementById('t-cat').value;
            const source = document.getElementById('t-source').value;
            const sort = document.getElementById('t-sort')?.value || 'date';
            const quality = document.getElementById('t-quality')?.value || '';
            const limit = document.getElementById('t-limit')?.value || '50';
            
            const url = new URL(window.location.origin + window.location.pathname);
            if (city) url.searchParams.set('city', city);
            if (q) url.searchParams.set('q', q);
            if (source) url.searchParams.set('source', source);
            if (sort) url.searchParams.set('sort', sort);
            if (quality) url.searchParams.set('quality', quality);
            
            // Preserve category parameter if present in current URL
            const currentUrl = new URL(window.location.href);
            const category = currentUrl.searchParams.get('category');
            if (category) url.searchParams.set('category', category);
            
            url.searchParams.set('limit', limit);
            url.searchParams.set('page', 1);
            window.loadLeads(url.toString(), true);
            connectStream();
        };

        window.setFilter = function(city, q) {
            document.getElementById('t-city').value = city;
            document.getElementById('t-cat').value = q;
            window.applyFilters();
        };

        window.clearFilters = function() {
            document.getElementById('t-city').value = '';
            document.getElementById('t-cat').value = '';
            document.getElementById('t-source').value = '';
            document.getElementById('t-sort').value = 'date';
            const ql = document.getElementById('t-quality');
            if (ql) ql.value = '';
            const ll = document.getElementById('t-limit');
            if (ll) ll.value = '50';
            
            const url = new URL(window.location.origin + window.location.pathname);
            url.searchParams.set('page', 1);
            url.searchParams.set('limit', '50');
            window.loadLeads(url.toString(), true);
            connectStream();
        };

        window.copyLead = function(text) {
            if (!text || text === '---') {
                window.showNotif('No data to copy');
                return;
            }
            navigator.clipboard.writeText(text).then(function() {
                window.showNotif('Copied to clipboard!');
            }).catch(function() {
                window.showNotif('Failed to copy');
            });
        };

        window.loadLeads = async function(url, pushState) {
            try {
                const tbody = document.getElementById('leads-tbody');
                if (tbody) {
                    tbody.innerHTML = '<tr class="skeleton-row"><td><div class="skeleton-cell skeleton-text"></div></td><td><div class="skeleton-cell skeleton-text"></div></td><td><div class="skeleton-cell skeleton-text"></div></td><td><div class="skeleton-cell skeleton-text"></div></td><td><div class="skeleton-cell skeleton-badge"></div></td><td><div class="skeleton-cell skeleton-text"></div></td><td><div class="skeleton-cell skeleton-badge"></div></td><td><div class="skeleton-cell skeleton-score"></div></td><td></td></tr>'.repeat(5);
                }
                const res = await fetch(url, { headers: { 'X-Requested-With': 'XMLHttpRequest' } });
                const data = await res.json();
                
                window.currentPage = data.page;
                window.totalPages = data.total_pages; if (data.limit) window.pageSize = data.limit;
                window.currentFilters = url;
                
                window.renderLeads(data.contacts);
                window.updatePaginationUI(data);
                window.updateStats(data.stats);
                window.updateFilterChips();
                
                if (pushState) {
                    history.pushState({page: data.page, url: url}, '', url);
                }
                
                document.querySelector('.glass-card')?.scrollIntoView({ behavior: 'smooth', block: 'start' });
            } catch (e) {
                console.error("AJAX Load Error:", e);
                window.showNotif('Failed to load data', 3000, true);
            }
        };

        window.escapeHtml = function(value) {
            if (value === null || value === undefined || value === '') return '---';
            return String(value)
                .replace(/&/g, '&amp;')
                .replace(/</g, '&lt;')
                .replace(/>/g, '&gt;')
                .replace(/"/g, '&quot;')
                .replace(/'/g, '&#39;');
        };

        window.escapeJsAttr = function(value) {
            if (value === null || value === undefined) return '';
            return String(value)
                .replace(/&/g, '&amp;')
                .replace(/"/g, '&quot;')
                .replace(/</g, '&lt;')
                .replace(/>/g, '&gt;');
        };

        window.updateStats = function(stats) {
            if (!stats) return;
            const totalEl = document.getElementById('stat-total');
            const phoneEl = document.getElementById('stat-phone');
            const emailEl = document.getElementById('stat-email');
            if (totalEl && stats.total !== undefined) totalEl.innerText = stats.total.toLocaleString();
            if (phoneEl && stats.phone !== undefined) phoneEl.innerText = stats.phone.toLocaleString();
            if (emailEl && stats.email !== undefined) emailEl.innerText = stats.email.toLocaleString();
        };

        window.updateFilterChips = function() {
            const chipsEl = document.getElementById('active-filter-chips');
            if (!chipsEl) return;
            const chips = [];
            const city = document.getElementById('t-city')?.value?.trim();
            const q = document.getElementById('t-cat')?.value?.trim();
            const source = document.getElementById('t-source')?.value;
            const quality = document.getElementById('t-quality')?.value;
            const sortEl = document.getElementById('t-sort');
            const sortText = sortEl?.options[sortEl.selectedIndex]?.text;
            if (city) chips.push({label: 'City: ' + city, onclick: function(){ document.getElementById('t-city').value=''; applyFilters(); }});
            if (q) chips.push({label: 'Search: ' + q, onclick: function(){ document.getElementById('t-cat').value=''; applyFilters(); }});
            if (source) chips.push({label: 'Source: ' + source, onclick: function(){ document.getElementById('t-source').value=''; applyFilters(); }});
            if (quality) chips.push({label: 'Quality: ' + quality, onclick: function(){ document.getElementById('t-quality').value=''; applyFilters(); }});
            var html = chips.map(function(c) {
                return '<span class="filter-chip" onclick="' + c.onclick.toString().replace(/"/g, '&quot;') + '">' + c.label + '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><line x1="18" y1="6" x2="6" y2="18"></line><line x1="6" y1="6" x2="18" y2="18"></line></svg></span>';
            }).join('') + '<span class="filter-sort-hint">sorted: ' + sortText + '</span>';
            chipsEl.innerHTML = html;
        };

        window.renderLeads = function(leads) {
            const tbody = document.getElementById('leads-tbody');
            if (!tbody) return;
            
            if (leads.length === 0) {
                tbody.innerHTML = '<tr><td colspan="9" style="text-align:center; padding:60px; color:var(--text-secondary);"><svg width="48" height="48" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1" style="opacity:0.3; margin-bottom:12px;"><circle cx="11" cy="11" r="8"></circle><line x1="21" y1="21" x2="16.65" y2="16.65"></line></svg><br>No records found matching filters</td></tr>';
                return;
            }

            tbody.innerHTML = leads.map(function(c, i) {
                const scoreColor = c.quality_score > 70 ? 'var(--accent-emerald)' : (c.quality_score > 40 ? 'var(--accent-blue)' : 'var(--accent-red)');
                const rowNum = (window.currentPage - 1) * window.pageSize + i + 1;
                const copyText = c.phone || c.email || '';
                const safeCopy = window.escapeJsAttr(copyText);
                return '<tr class="lead-row">' +
                    '<td><div class="row-checkbox" onclick="toggleRow(this)"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="3"><polyline points="20 6 9 17 4 12"></polyline></svg></div></td>' +
                    '<td style="color:var(--text-muted); font-size:11px;">' + rowNum + '</td>' +
                    '<td class="cell-truncate" style="font-weight:700; font-family:\\'Outfit\\',sans-serif; color:#fff;">' + window.escapeHtml(c.name) + '</td>' +
                    '<td class="mono cell-truncate" style="font-size:12px;">' + window.escapeHtml(c.phone) + '</td>' +
                    '<td class="mono cell-truncate" style="color:var(--accent-blue); font-size:11px;">' + window.escapeHtml(c.email) + '</td>' +
                    '<td class="cell-truncate" style="font-size:12px; font-weight:500;">' + window.escapeHtml(c.category) + '</td>' +
                    '<td style="font-size:12px; color:var(--text-secondary);">' + window.escapeHtml(c.city) + '</td>' +
                    '<td><span class="badge badge-src">' + window.escapeHtml(c.source) + '</span></td>' +
                    '<td>' +
                        '<div class="score-wrapper">' +
                            '<div class="score-bar"><div class="score-fill" style="width:' + c.quality_score + '%; background:' + scoreColor + ';"></div></div>' +
                            '<span class="mono score-value">' + c.quality_score + '%</span>' +
                        '</div>' +
                    '</td>' +
                    '<td>' +
                        '<div style="display:flex; gap:4px;">' +
                            '<button class="action-btn" title="Copy" data-copy="' + safeCopy + '" onclick="copyLead(this.dataset.copy)">' +
                                '<svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="9" y="9" width="13" height="13" rx="2"></rect><path d="M5 15H4a2 2 0 0 1-2-2V4a2 2 0 0 1 2-2h9a2 2 0 0 1 2 2v1"></path></svg>' +
                            '</button>' +
                        '</div>' +
                    '</td>' +
                '</tr>';
            }).join('');
        };

window.updatePaginationUI = function(data) {
            var btnContainer = document.getElementById('pagination-btns-container');
            if (!btnContainer) return;

            var html = '';
            var isFirst = data.page <= 1;
            var isLast = data.page >= data.total_pages;

            html += '<button class="pagination-btn icon-btn" onclick="goToPage(1)" ' + (isFirst ? 'disabled' : '') + ' title="First">';
            html += '<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polyline points="11 17 6 12 11 7"></polyline><polyline points="18 17 13 12 18 7"></polyline></svg>';
            html += '</button>';
            html += '<button class="pagination-btn" onclick="changePage(-1)" ' + (isFirst ? 'disabled' : '') + '>Prev</button>';

            var start_p = Math.max(1, data.page - 2);
            var end_p = Math.min(data.total_pages, start_p + 4);
            start_p = Math.max(1, end_p - 4);

            for (var i = start_p; i <= end_p; i++) {
                html += '<button class="pagination-btn ' + (i === data.page ? 'active' : '') + '" onclick="goToPage(' + i + ')">' + i + '</button>';
            }

            html += '<button class="pagination-btn" onclick="changePage(1)" ' + (isLast ? 'disabled' : '') + '>Next</button>';
            html += '<button class="pagination-btn icon-btn" onclick="goToPage(' + data.total_pages + ')" ' + (isLast ? 'disabled' : '') + ' title="Last">';
            html += '<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polyline points="13 17 18 12 13 7"></polyline><polyline points="6 17 11 12 6 7"></polyline></svg>';
            html += '</button>';

            btnContainer.innerHTML = html;

            // Update info
            var infoEl = document.querySelector('.pagination-info');
            if (infoEl) {
                var totalText = data.filtered_total !== undefined ? data.filtered_total.toLocaleString() : (data.total_pages * data.contacts.length).toLocaleString();
                var startRange = data.contacts.length > 0 ? (data.page - 1) * (data.limit || window.pageSize) + 1 : 0;
                var endRange = data.contacts.length > 0 ? startRange + data.contacts.length - 1 : 0;
                
                infoEl.innerHTML = '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"></path><circle cx="9" cy="7" r="4"></circle><path d="M23 21v-2a4 4 0 0 0-3-3.87"></path><path d="M16 3.13a4 4 0 0 1 0 7.75"></path></svg>' +
                    '<span><span>' + startRange + '-' + endRange + '</span> of <span class="total-count">' + totalText + '</span> leads</span>' +
                    '<span style="color:var(--border-muted);">|</span>' +
                    '<span>Page <span>' + data.page + '</span> of <span>' + data.total_pages + '</span></span>';
            }
        };

        // Handle Browser Back/Forward
        window.addEventListener('popstate', function(event) {
            if (event.state && event.state.url) {
                window.loadLeads(event.state.url, false);
            }
        });
        
        // Initialize history state on load
        if (typeof history.replaceState === 'function') {
            history.replaceState({page: window.currentPage, url: window.location.href}, '', window.location.href);
        }

        window.startCollection = async function(autoMode) {
            const city = document.getElementById('t-city').value;
            const q = document.getElementById('t-cat').value;
            const source = document.getElementById('t-source').value;
            const btn = document.getElementById('start-btn');
            const autoBtn = document.getElementById('auto-btn');
            const shouldAuto = !!autoMode || (!city && !q && !source);
            
            if (shouldAuto) {
                if (autoBtn) autoBtn.style.display = 'none';
                const stopBtn = document.getElementById('stop-btn');
                if (stopBtn) stopBtn.style.display = 'inline-flex';
            }

            if (btn) {
                btn.disabled = true;
                btn.innerHTML = '<span class="pulse">COLLECTING...</span>';
            }
            if (autoBtn) autoBtn.disabled = true;
            
            try {
                const res = await fetch('/api/trigger/scrape', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({city, category: q, source, auto: shouldAuto})
                });
                const data = await res.json();
                window.showNotif(data.message);
            } catch (e) {
                window.showNotif('Failed to trigger collection');
                if (btn) {
                    btn.disabled = false;
                    btn.innerText = 'Start Collection';
                }
                if (autoBtn) autoBtn.disabled = false;
            }
        };

        window.setTemplate = function(city, q, src) {
            document.getElementById('t-city').value = city;
            document.getElementById('t-cat').value = q;
            document.getElementById('t-source').value = src;
            window.applyFilters();
        };

        window.exportData = function(fmt, btn) {
            const q = document.getElementById('t-cat')?.value || "";
            const city = document.getElementById('t-city')?.value || "";
            const src = document.getElementById('t-source')?.value || "";
            
            if (btn) {
                const originalHTML = btn.innerHTML;
                btn.disabled = true;
                btn.innerHTML = '<div class="spinner-sm"></div>';
                setTimeout(() => {
                    btn.disabled = false;
                    btn.innerHTML = originalHTML;
                }, 5000);
            }
            
            const baseUrl = window.location.origin;
            let url = baseUrl + "/export/" + fmt;
            const params = new URLSearchParams();
            if (city) params.set('city', city);
            if (q) params.set('q', q);
            if (src) params.set('source', src);
            
            // Preserve category parameter if present in current URL
            const currentUrl = new URL(window.location.href);
            const category = currentUrl.searchParams.get('category');
            if (category) params.set('category', category);
            
            if (window.isSchoolDashboard) {
                params.set('schools_only', 'true');
            } else {
                params.set('financial_only', 'true');
            }
            
            const queryString = params.toString();
            if (queryString) url += "?" + queryString;
            window.location.assign(url);
        };

        window.exportAllData = function(fmt, btn) {
            if (btn) {
                const originalHTML = btn.innerHTML;
                btn.disabled = true;
                btn.innerHTML = '<div class="spinner-sm"></div>';
                setTimeout(() => {
                    btn.disabled = false;
                    btn.innerHTML = originalHTML;
                }, 5000);
            }
            const q = document.getElementById('t-cat')?.value || "";
            const city = document.getElementById('t-city')?.value || "";
            const src = document.getElementById('t-source')?.value || "";
            const quality = document.getElementById('t-quality')?.value || "";
            const baseUrl = window.location.origin;
            let url = baseUrl + "/export/" + fmt + "?all=true";
            if (window.isSchoolDashboard) {
                url += "&schools_only=true";
            } else {
                url += "&financial_only=true";
            }
            if (city) url += "&city=" + encodeURIComponent(city);
            if (q) url += "&q=" + encodeURIComponent(q);
            if (src) url += "&source=" + encodeURIComponent(src);
            if (quality) url += "&quality=" + encodeURIComponent(quality);
            window.location.assign(url);
        };

        window.exportCategory = function(category, fmt, btn) {
            if (btn) {
                const originalHTML = btn.innerHTML;
                btn.disabled = true;
                btn.innerHTML = '<div class="spinner-sm"></div>';
                setTimeout(() => {
                    btn.disabled = false;
                    btn.innerHTML = originalHTML;
                }, 5000);
            }
            let url = window.location.origin + "/export/" + fmt + "?all=true&category=" + encodeURIComponent(category);
            if (window.isSchoolDashboard) {
                url += "&schools_only=true";
            } else {
                url += "&financial_only=true";
            }
            window.location.assign(url);
        };

        window.stopScraping = async function() {
            try {
                const res = await fetch('/api/trigger/stop', { method: 'POST' });
                const data = await res.json();
                window.showNotif(data.message);
                document.getElementById('stop-btn').style.display = 'none';
                document.getElementById('auto-btn').style.display = 'inline-flex';
            } catch (e) {
                window.showNotif('Failed to send stop signal');
            }
        };

        window.cleanup = async function() {
            window.showNotif('Cleaning started...');
            try {
                const res = await fetch('/api/cleanup/deep', {method: 'POST'});
                const data = await res.json();
                window.showNotif('Done: ' + data.deleted + ' deleted');
            } catch(e) { window.showNotif('Cleanup failed'); }
        };

        window.updateQuality = async function() {
            window.showNotif('Quality audit started...');
            try {
                const res = await fetch('/api/cleanup/quality', {method: 'POST'});
                const data = await res.json();
                window.showNotif('Audited ' + data.updated + ' records');
            } catch(e) { window.showNotif('Audit failed'); }
        };

        // Live Telemetry Stream
        let evtSource = null;
        function connectStream() {
            if (evtSource) evtSource.close();
            let streamUrl = "/api/stream/stats";
            const params = new URLSearchParams();
            if (window.isSchoolDashboard) params.set('schools_only', 'true');
            const q = document.getElementById('t-cat')?.value;
            const city = document.getElementById('t-city')?.value;
            const src = document.getElementById('t-source')?.value;
            const quality = document.getElementById('t-quality')?.value;
            if (q) params.set('q', q);
            if (city) params.set('city', city);
            if (src) params.set('source', src);
            if (quality) params.set('quality', quality);
            const qs = params.toString();
            if (qs) streamUrl += "?" + qs;
            evtSource = new EventSource(streamUrl);
            evtSource.onmessage = function(event) {
            const data = JSON.parse(event.data);
            if (document.getElementById('stat-total')) {
                const el = document.getElementById('stat-total');
                el.innerText = data.total;
                el.classList.add('animate');
                setTimeout(() => el.classList.remove('animate'), 400);
            }
            if (document.getElementById('stat-phone')) {
                const el = document.getElementById('stat-phone');
                el.innerText = data.with_phone;
                el.classList.add('animate');
                setTimeout(() => el.classList.remove('animate'), 400);
            }
            if (document.getElementById('stat-email')) {
                const el = document.getElementById('stat-email');
                el.innerText = data.with_email;
                el.classList.add('animate');
                setTimeout(() => el.classList.remove('animate'), 400);
            }
            if (document.getElementById('last-update')) document.getElementById('last-update').innerText = new Date().toLocaleTimeString();
                  const status = data.scraper_status;
            const statusEl = document.getElementById('live-status');
            const progWrap = document.getElementById('prog-wrap');
            const progBar = document.getElementById('prog-bar');
            
            const startBtn = document.getElementById('start-btn');
            const autoBtn = document.getElementById('auto-btn');
            const stopBtn = document.getElementById('stop-btn');
            
            const globalStartBtn = document.getElementById('global-start-btn');
            const globalStopBtn = document.getElementById('global-stop-btn');

            if (status && status.running) {
                const cleanMsg = (status.message || 'SCRAPING').toUpperCase();
                if(statusEl) { 
                    statusEl.innerText = cleanMsg; 
                    statusEl.style.color = 'var(--accent-emerald)'; 
                }
                if(progWrap) progWrap.style.display = 'block';
                if(progBar) progBar.style.width = (status.stats && status.stats.progress ? status.stats.progress : 100) + '%';
                
                // Update local filter buttons
                if (startBtn) {
                    startBtn.disabled = true;
                    startBtn.innerHTML = '<span class="pulse">COLLECTING...</span>';
                }
                if (autoBtn) autoBtn.disabled = true;
                if (stopBtn) stopBtn.style.display = 'inline-flex';

                // Update unified global HUD buttons
                if (globalStartBtn) {
                    globalStartBtn.disabled = true;
                    globalStartBtn.style.opacity = '0.5';
                    globalStartBtn.style.cursor = 'not-allowed';
                    globalStartBtn.innerHTML = '<span class="pulse" style="display:flex;align-items:center;gap:4px;"><span class="status-dot-pulse"></span>COLLECTING...</span>';
                }
                if (globalStopBtn) {
                    globalStopBtn.disabled = false;
                    globalStopBtn.style.opacity = '1';
                    globalStopBtn.style.cursor = 'pointer';
                }
            } else {
                if(statusEl) { 
                    statusEl.innerText = 'IDLE / ONLINE'; 
                    statusEl.style.color = 'var(--text-secondary)'; 
                }
                if(progWrap) progWrap.style.display = 'none';
                
                // Update local filter buttons
                if (startBtn) {
                    startBtn.disabled = false;
                    startBtn.innerHTML = '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polygon points="5 3 19 12 5 21 5 3"></polygon></svg> Start Collection';
                }
                if (autoBtn) autoBtn.disabled = false;
                if (stopBtn) stopBtn.style.display = 'none';

                // Update unified global HUD buttons
                if (globalStartBtn) {
                    globalStartBtn.disabled = false;
                    globalStartBtn.style.opacity = '1';
                    globalStartBtn.style.cursor = 'pointer';
                    globalStartBtn.innerHTML = '<svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="3"><polygon points="5 3 19 12 5 21 5 3"></polygon></svg> START SCRAPING';
                }
                if (globalStopBtn) {
                    globalStopBtn.disabled = true;
                    globalStopBtn.style.opacity = '0.5';
                    globalStopBtn.style.cursor = 'not-allowed';
                }
            }

            // Activity logs hidden per user request
            
            const badge = document.getElementById('status-badge');
            if (badge) {
                if (status && status.running) {
                    badge.innerText = 'SCRAPING'; badge.style.color = 'var(--accent-emerald)';
                } else {
                    badge.innerText = 'ONLINE'; badge.style.color = 'var(--text-secondary)';
                }
            }
        };
        }

        connectStream();  // initial connection with current filters

        // Live Feed SSE (combined with telemetry above)
        
        let sourceChart, categoryChart, trendChart;
        async function initCharts() {
            const chartColors = ['#10b981', '#3b82f6', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899'];
            const fontOpt = { color: '#94a3b8', font: { family: 'Inter', size: 10 } };
            const baseOpt = { 
                responsive: true, 
                maintainAspectRatio: false, 
                animation: { duration: 800, easing: 'easeOutQuart' },
                plugins: {
                    legend: {
                        labels: {
                            color: '#94a3b8',
                            font: { family: 'Inter', size: 10, weight: '500' },
                            padding: 10,
                            usePointStyle: true
                        }
                    }
                }
            };

            const srcEl = document.getElementById('sourceChart');
            if (srcEl) {
                sourceChart = new Chart(srcEl, {
                    type: 'doughnut',
                    data: { labels: [], datasets: [{ data: [], backgroundColor: chartColors, borderColor: 'rgba(0,0,0,0.2)', borderWidth: 2 }] },
                    options: { 
                        ...baseOpt, 
                        plugins: { ...baseOpt.plugins, legend: { ...baseOpt.plugins.legend, position: 'bottom' } },
                        cutout: '65%' 
                    }
                });
            }

            const catEl = document.getElementById('categoryChart');
            if (catEl) {
                categoryChart = new Chart(catEl, {
                    type: 'bar',
                    data: { labels: [], datasets: [{ data: [], backgroundColor: 'rgba(16, 185, 129, 0.7)', borderRadius: 6 }] },
                    options: { 
                        ...baseOpt, 
                        plugins: { ...baseOpt.plugins, legend: { display: false } }, 
                        scales: { 
                            x: { ticks: fontOpt, grid: { display: false } }, 
                            y: { ticks: fontOpt, grid: { color: 'rgba(255,255,255,0.03)' } }
                        }
                    }
                });
            }

            const trendEl = document.getElementById('trendChart');
            if (trendEl) {
                trendChart = new Chart(trendEl, {
                    type: 'line',
                    data: { labels: [], datasets: [{ 
                        data: [], 
                        borderColor: '#3b82f6', 
                        backgroundColor: 'rgba(59, 130, 246, 0.1)',
                        fill: true,
                        tension: 0.4,
                        borderWidth: 3,
                        pointRadius: 0
                    }] },
                    options: { 
                        ...baseOpt, 
                        plugins: { ...baseOpt.plugins, legend: { display: false } }, 
                        scales: { 
                            x: { ticks: fontOpt, grid: { display: false } }, 
                            y: { ticks: fontOpt, grid: { color: 'rgba(255,255,255,0.03)' } }
                        }
                    }
                });
            }
            
            const growthEl = document.getElementById('growthChart');
            if (growthEl) {
                window.growthChart = new Chart(growthEl, {
                    type: 'line',
                    data: { labels: [], datasets: [{ 
                        data: [], 
                        borderColor: '#10b981', 
                        backgroundColor: 'rgba(16, 185, 129, 0.15)',
                        fill: true,
                        tension: 0.4,
                        borderWidth: 2,
                        pointRadius: 0
                    }] },
                    options: { 
                        ...baseOpt, 
                        plugins: { ...baseOpt.plugins, legend: { display: false } }, 
                        scales: { 
                            x: { ticks: fontOpt, grid: { display: false } }, 
                            y: { ticks: fontOpt, grid: { color: 'rgba(255,255,255,0.03)' } }
                        }
                    }
                });
            }

            refreshCharts();
            setInterval(refreshCharts, 30000);
        }

        async function refreshCharts() {
            try {
                let url = '/api/stats/charts';
                if (window.isSchoolDashboard) {
                    url += '?schools_only=true';
                }
                const response = await fetch(url);
                const stats = await response.json();
                if (!stats.sources) return;
                
                if (sourceChart) {
                    sourceChart.data.labels = stats.sources.map(function(s) { return s.source; });
                    sourceChart.data.datasets[0].data = stats.sources.map(function(s) { return s.count; });
                    sourceChart.update();
                }
                
                if (categoryChart) {
                    categoryChart.data.labels = stats.categories.slice(0,5).map(function(c) { return c.category; });
                    categoryChart.data.datasets[0].data = stats.categories.slice(0,5).map(function(c) { return c.count; });
                    categoryChart.update();
                }
                
                if (trendChart) {
                    trendChart.data.labels = stats.trend.map(function(t) { return t.date; });
                    trendChart.data.datasets[0].data = stats.trend.map(function(t) { return t.count; });
                    trendChart.update();
                }
                
                if (window.growthChart) {
                    window.growthChart.data.labels = stats.trend.map(function(t) { return t.date; });
                    window.growthChart.data.datasets[0].data = stats.trend.map(function(t) { return t.count; });
                    window.growthChart.update();
                }
            } catch(e) { console.log('Chart error:', e); }
        }

        window.openMaintenance = async function() {
            if (!confirm("Run system-wide category normalization? This will fix duplicate charts by merging similar categories.")) return;
            
            showNotif("Starting database maintenance...");
            try {
                const resp = await fetch('/api/maintenance/normalize', { method: 'POST' });
                const res = await resp.json();
                if (res.success) {
                    showNotif(`Success! Normalized ${res.category_normalized} entries.`);
                    refreshCharts();
                } else {
                    showNotif("Error: " + res.error);
                }
            } catch(e) {
                showNotif("Maintenance failed: " + e);
            }
        };

        if (document.getElementById('sourceChart')) {
            initCharts();
        }
        
        // Keyboard shortcuts
        document.addEventListener('keydown', function(e) {
            if (e.target.tagName === 'INPUT' || e.target.tagName === 'SELECT') return;
            
            if (e.key === 'ArrowLeft' && !e.ctrlKey) {
                e.preventDefault();
                goToPage(currentPage - 1);
            }
            if (e.key === 'ArrowRight' && !e.ctrlKey) {
                e.preventDefault();
                goToPage(currentPage + 1);
            }
        });
        
        // Enter key on search inputs
        document.getElementById('t-cat')?.addEventListener('keydown', function(e) {
            if (e.key === 'Enter') applyFilters();
        });
        document.getElementById('t-city')?.addEventListener('keydown', function(e) {
            if (e.key === 'Enter') applyFilters();
        });
        
        // Theme Toggle
        window.toggleTheme = function() {
            const html = document.documentElement;
            const isLight = html.getAttribute('data-theme') === 'light';
            const newTheme = isLight ? 'dark' : 'light';
            html.setAttribute('data-theme', newTheme);
            localStorage.setItem('theme', newTheme);
            
            document.getElementById('theme-icon-sun').style.display = isLight ? 'block' : 'none';
            document.getElementById('theme-icon-moon').style.display = isLight ? 'none' : 'block';
        };
        
        // Restore theme from localStorage
        const savedTheme = localStorage.getItem('theme');
        if (savedTheme) {
            document.documentElement.setAttribute('data-theme', savedTheme);
            if (savedTheme === 'light') {
                document.getElementById('theme-icon-sun').style.display = 'none';
                document.getElementById('theme-icon-moon').style.display = 'block';
            }
        }
        
        // Row Checkbox Selection
        window.toggleSelectAll = function(el) {
            const checked = !el.classList.contains('checked');
            el.classList.toggle('checked', checked);
            document.querySelectorAll('#leads-tbody .row-checkbox').forEach(cb => {
                cb.classList.toggle('checked', checked);
            });
        };
        
        window.toggleRow = function(el) {
            el.classList.toggle('checked');
        };
        
        // Direct Scraping Functions (No Proxy)
        window.startDirectScrape = async function(source) {
            const city = document.getElementById('t-city')?.value || "";
            const category = document.getElementById('t-cat')?.value || "";
            
            showNotif(`Starting direct scrape for ${source}...`);
            
            try {
                const params = new URLSearchParams();
                params.set('source', source);
                if (city) params.set('city', city);
                if (category) params.set('category', category);
                
                const res = await fetch('/api/trigger/direct-scrape?' + params.toString(), {
                    method: 'POST'
                });
                const data = await res.json();
                
                if (data.error) {
                    showNotif(data.error, 4000, true);
                } else {
                    showNotif(data.message);
                }
            } catch(e) {
                showNotif('Failed to start direct scrape', 3000, true);
            }
        };
        
        window.startGovBatch = async function() {
            showNotif('Starting government sites batch scrape...');
            
            try {
                const res = await fetch('/api/trigger/direct-gov-batch', {
                    method: 'POST'
                });
                const data = await res.json();
                
                if (data.error) {
                    showNotif(data.error, 4000, true);
                } else {
                    showNotif(data.message);
                }
            } catch(e) {
                showNotif('Failed to start gov batch', 3000, true);
            }
        };

        window.setSchoolZone = function(zone) {
            document.getElementById('school-zone').value = zone;
        };

        window.scrapeSchools = async function() {
            const zone = document.getElementById('school-zone')?.value || '';
            const statusEl = document.getElementById('school-scrape-status');
            statusEl.textContent = 'Starting school scrape for ' + (zone || 'all India') + '...';
            statusEl.style.color = 'var(--accent-emerald, #10b981)';
            
            try {
                const params = new URLSearchParams();
                params.set('source', 'SCHOOL');
                if (zone) params.set('city', zone);
                params.set('category', 'Schools');
                
                const res = await fetch('/api/trigger/direct-scrape?' + params.toString(), {
                    method: 'POST'
                });
                const data = await res.json();
                
                if (data.error) {
                    statusEl.textContent = 'Error: ' + data.error;
                    statusEl.style.color = 'var(--accent-red, #ef4444)';
                    showNotif(data.error, 4000, true);
                } else {
                    statusEl.textContent = 'Scraping ' + (zone || 'all India') + ' — task queued! Check logs for progress.';
                    showNotif('School scraper started for ' + (zone || 'all India'));
                }
            } catch(e) {
                statusEl.textContent = 'Failed to start school scrape';
                statusEl.style.color = 'var(--accent-red, #ef4444)';
                showNotif('Failed to start school scrape', 3000, true);
            }
        };

        {% if is_school_dashboard %}
        window.pollSchoolStatus = async function() {
            const statusEl = document.getElementById('school-scrape-status');
            if (!statusEl) return;
            try {
                const res = await fetch('/api/status');
                const data = await res.json();
                if (data && data.message) {
                    const msg = data.message || '';
                    if (msg.toLowerCase().includes('school')) {
                        statusEl.textContent = msg.substring(0, 60);
                        statusEl.style.color = 'var(--accent-emerald, #10b981)';
                    } else if (data.is_running) {
                        statusEl.textContent = 'System busy: ' + msg.substring(0, 40);
                        statusEl.style.color = 'var(--accent-emerald, #10b981)';
                    } else if (statusEl.textContent === 'Ready' || statusEl.textContent.includes('Ready')) {
                        // keep as is
                    }
                }
            } catch(e) {}
        };
        setInterval(window.pollSchoolStatus, 5000);
        {% endif %}

        window.startScrapeGlobal = async function() {
            const isSchool = {% if is_school_dashboard %}true{% else %}false{% endif %};
            if (isSchool) {
                if (typeof window.scrapeSchools === 'function') {
                    await window.scrapeSchools();
                } else {
                    await startDirectScrape('SCHOOL');
                }
            } else {
                showNotif('Activating continuous AutoPilot scraping...');
                try {
                    const res = await fetch('/api/trigger/scrape?auto=true', { method: 'POST' });
                    const data = await res.json();
                    showNotif(data.message);
                } catch(e) {
                    showNotif('Failed to activate AutoPilot', 3000, true);
                }
            }
        };
        
        window.stopScrapeGlobal = async function() {
            showNotif('Sending STOP signal to active scraper...');
            
            // Immediate UI feedback
            const statusEl = document.getElementById('live-status');
            if (statusEl) {
                statusEl.innerText = 'STOPPING...';
                statusEl.style.color = 'var(--accent-red)';
            }
            const globalStopBtn = document.getElementById('global-stop-btn');
            if (globalStopBtn) {
                globalStopBtn.disabled = true;
                globalStopBtn.style.opacity = '0.5';
            }
            
            try {
                const res = await fetch('/api/trigger/stop', { method: 'POST' });
                const data = await res.json();
                showNotif(data.message);
                
                // Reset local filter buttons
                const startBtn = document.getElementById('start-btn');
                if (startBtn) {
                    startBtn.disabled = false;
                    startBtn.innerHTML = '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polygon points="5 3 19 12 5 21 5 3"></polygon></svg> Start Collection';
                }
                const autoBtn = document.getElementById('auto-btn');
                if (autoBtn) autoBtn.disabled = false;
                const stopBtn = document.getElementById('stop-btn');
                if (stopBtn) stopBtn.style.display = 'none';
                
                const schoolStatus = document.getElementById('school-scrape-status');
                if (schoolStatus) {
                    schoolStatus.textContent = 'Ready';
                    schoolStatus.style.color = 'var(--text-muted)';
                }
            } catch(e) {
                showNotif('Failed to stop scraping', 3000, true);
            }
        };

        /* ===== AUTO-REFRESH ===== */

        // Keyboard shortcuts
        document.addEventListener('keydown', function(e) {
            if (e.target.tagName === 'INPUT' || e.target.tagName === 'SELECT' || e.target.tagName === 'TEXTAREA') {
                if (e.key === 'Enter' && (e.target.id === 't-cat' || e.target.id === 't-city')) {
                    e.preventDefault();
                    applyFilters();
                }
                return;
            }
            switch(e.key) {
                case 'r': case 'R': applyFilters(); break;
                case 'Escape': clearFilters(); break;
            }
        });

        // Init
    </script>
</body>
</html>
"""
def render_dashboard_portal(is_school_dashboard=False):
    try:
        config = load_config()
        scraper_cfg = config.get("scraper", {})
        page_size = int(
            os.environ.get(
                "DASHBOARD_PAGE_SIZE", scraper_cfg.get("dashboard_page_size", 50)
            )
        )

        page = request.args.get("page", 1, type=int)
        limit = request.args.get("limit", page_size, type=int)
        limit = max(1, min(limit, 500))

        search_query = request.args.get("q", "")
        selected_city = request.args.get("city", "")
        selected_category = request.args.get("category", "")
        selected_source = request.args.get("source", "")
        selected_quality = request.args.get("quality", "")
        sort_by = request.args.get("sort", "date")

        conn = get_db()
        cur = conn.cursor()

        # Sort mapping
        sort_map = {
            "date": "scraped_at DESC",
            "date_asc": "scraped_at ASC",
            "name": "LOWER(name) ASC",
            "name_desc": "LOWER(name) DESC",
            "city": "LOWER(city) ASC",
            "city_desc": "LOWER(city) DESC",
            "source": "LOWER(source) ASC",
            "source_desc": "LOWER(source) DESC",
            "score": "quality_score DESC",
            "score_asc": "quality_score ASC",
            "category": "LOWER(category) ASC",
            "category_desc": "LOWER(category) DESC",
            "phone": "phone ASC",
            "phone_desc": "phone DESC",
            "email": "email ASC",
            "email_desc": "email DESC",
        }
        order_by = sort_map.get(sort_by, "scraped_at DESC")

        where_sql, params = build_contact_filters(
            search_query,
            selected_city,
            selected_category,
            selected_source,
            selected_quality,
            exclude_schools=not is_school_dashboard,
            only_schools=is_school_dashboard
        )

        ph = db_placeholder()
        like_op = "LIKE" if USE_SQLITE else "ILIKE"
        if is_school_dashboard:
            where_clause = f"(LOWER(category) {like_op} {ph} OR LOWER(source) IN ('npsc', 'bsai', 'aisa'))"
        else:
            where_clause = f"(LOWER(category) NOT {like_op} {ph} AND LOWER(source) NOT IN ('npsc', 'bsai', 'aisa'))"
        params_stat = ["%school%"]

        # Get total count (unfiltered for this portal context)
        cur.execute(f"SELECT COUNT(*) as cnt FROM contacts WHERE {where_clause}", params_stat)
        total = cur.fetchone()["cnt"]

        # Get filtered count
        count_sql = f"SELECT COUNT(*) as cnt FROM contacts WHERE {where_sql}"
        cur.execute(count_sql, params)
        filtered_total = cur.fetchone()["cnt"]

        # Final Page Calculation
        total_pages = (filtered_total + limit - 1) // limit if filtered_total > 0 else 1
        
        # Clamp and validate current page
        if page < 1: page = 1
        if page > total_pages: page = total_pages
        
        offset = (page - 1) * limit

        query_sql = f"SELECT id, name, phone, email, city, source, category, quality_tier, quality_score, scraped_at FROM contacts WHERE {where_sql} ORDER BY {order_by} LIMIT {ph} OFFSET {ph}"
        cur.execute(query_sql, params + [limit, offset])
        contacts = cur.fetchall()

        # Get unique values for filter dropdowns (CACHED)
        cities = get_cached_filter(
            f"cities_{'school' if is_school_dashboard else 'fin'}",
            f"SELECT DISTINCT city FROM contacts WHERE {where_clause} AND city IS NOT NULL AND city <> '' ORDER BY city",
            cur,
            params_stat
        )
        categories = get_cached_filter(
            f"categories_{'school' if is_school_dashboard else 'fin'}",
            f"SELECT DISTINCT category FROM contacts WHERE {where_clause} AND category IS NOT NULL AND category <> '' ORDER BY category",
            cur,
            params_stat
        )
        sources = get_cached_filter(
            f"sources_{'school' if is_school_dashboard else 'fin'}",
            f"SELECT DISTINCT source FROM contacts WHERE {where_clause} AND source IS NOT NULL AND source <> '' ORDER BY source",
            cur,
            params_stat
        )

        # Stats reflect active filters (city, category, source, search, quality)
        if USE_SQLITE:
            cur.execute(f"""
                SELECT 
                    SUM(CASE WHEN phone_clean IS NOT NULL AND phone_clean <> '' THEN 1 ELSE 0 END) as with_phone,
                    SUM(CASE WHEN email IS NOT NULL AND email <> '' THEN 1 ELSE 0 END) as with_email,
                    COUNT(DISTINCT city) as city_count,
                    SUM(CASE WHEN LOWER(quality_tier) = 'high' THEN 1 ELSE 0 END) as q_high,
                    SUM(CASE WHEN LOWER(quality_tier) = 'medium' THEN 1 ELSE 0 END) as q_medium,
                    SUM(CASE WHEN LOWER(quality_tier) = 'low' THEN 1 ELSE 0 END) as q_low,
                    AVG(quality_score) as avg_score
                FROM contacts
                WHERE {where_sql}
            """, params)
        else:
            cur.execute(f"""
                SELECT 
                    COUNT(*) FILTER (WHERE phone_clean IS NOT NULL AND phone_clean <> '') as with_phone,
                    COUNT(*) FILTER (WHERE email IS NOT NULL AND email <> '') as with_email,
                    COUNT(DISTINCT city) as city_count,
                    COUNT(*) FILTER (WHERE LOWER(quality_tier) = 'high') as q_high,
                    COUNT(*) FILTER (WHERE LOWER(quality_tier) = 'medium') as q_medium,
                    COUNT(*) FILTER (WHERE LOWER(quality_tier) = 'low') as q_low,
                    AVG(quality_score) as avg_score
                FROM contacts
                WHERE {where_sql}
            """, params)
        stats_row = cur.fetchone()
        if stats_row:
            stats_row = dict(stats_row)
        
        cur.execute(f"SELECT source, COUNT(*) as c FROM contacts WHERE {where_sql} GROUP BY source", params)
        by_source = {r["source"]: r["c"] for r in cur.fetchall()}
        cur.execute(f"SELECT category, COUNT(*) as c FROM contacts WHERE {where_sql} GROUP BY category", params)
        by_cat = {r["category"]: r["c"] for r in cur.fetchall()}
        cur.close()
        conn.close()
    except Exception as e:
        logger.error(f"Database error in portal: {e}")
        contacts, total, filtered_total, stats_row = [], 0, 0, {}
        by_source, by_cat, total_pages, page = {}, {}, 1, 1
        cities, categories, sources = [], [], []
        selected_city = selected_category = selected_source = ""
        selected_quality = ""
        search_query = ""
        sort_by = "date"
        limit = page_size

    # Support AJAX / JSON response for flicker-free pagination
    if request.args.get("format") == "json" or request.headers.get("X-Requested-With") == "XMLHttpRequest":
        leads_list = []
        for c in contacts:
            leads_list.append({
                "id": c["id"],
                "name": c["name"],
                "phone": c["phone"] or "---",
                "email": c["email"] or "---",
                "city": c["city"] or "---",
                "source": c["source"] or "---",
                "category": c["category"] or "---",
                "quality_score": c["quality_score"] or 0
            })
        return jsonify({
            "contacts": leads_list,
            "page": page, "limit": limit,
            "total_pages": total_pages,
            "filtered_total": filtered_total,
            "stats": {
                "total": total,
                "phone": stats_row.get("with_phone", 0) if stats_row else 0,
                "email": stats_row.get("with_email", 0) if stats_row else 0,
            }
        })

    return render_template_string(
        HTML,
        is_school_dashboard=is_school_dashboard,
        contacts=contacts,
        s={
            "total": total,
            "phone": stats_row.get("with_phone", 0) if stats_row else 0,
            "email": stats_row.get("with_email", 0) if stats_row else 0,
            "cities": stats_row.get("city_count", 0) if stats_row else 0,
            "filtered_total": filtered_total,
            "quality_high": stats_row.get("q_high", 0) if stats_row else 0,
            "quality_medium": stats_row.get("q_medium", 0) if stats_row else 0,
            "quality_low": stats_row.get("q_low", 0) if stats_row else 0,
            "avg_quality": round(stats_row.get("avg_score", 0) or 0, 1) if stats_row else 0,
            "with_phone_pct": round((stats_row.get("with_phone", 0) / total * 100) if total > 0 else 0, 1),
            "with_email_pct": round((stats_row.get("with_email", 0) / total * 100) if total > 0 else 0, 1),
        },
        by_source=by_source,
        by_cat=by_cat,
        page=page,
        total_pages=total_pages,
        cities_default=config.get("cities", []),
        categories_default=config.get("categories", []),
        cities=cities,
        categories=categories,
        sources=sources,
        selected_city=selected_city,
        selected_category=selected_category,
        selected_source=selected_source,
        selected_quality=selected_quality,
        search_query=search_query,
        sort_by=sort_by,
        limit=limit,
    )


@app.route("/")
def index():
    return render_dashboard_portal(is_school_dashboard=False)


@app.route("/schools")
def schools_index():
    return render_dashboard_portal(is_school_dashboard=True)


@app.route("/api/status")
def get_status():
    def db_status():
        try:
            conn = get_db()
            cur = conn.cursor()
            if USE_SQLITE:
                cur.execute("SELECT value FROM system_status WHERE key = ?", ("scraper_status",))
            else:
                cur.execute("SELECT value FROM system_status WHERE key = %s", ("scraper_status",))
            row = cur.fetchone()
            cur.close()
            conn.close()
            if row:
                return json.loads(row["value"])
        except Exception:
            pass
        return None

    try:
        if redis_client:
            status = redis_client.get("scraper_status")
            if status:
                return Response(status, mimetype="application/json")
    except Exception:
        pass

    fallback = db_status()
    if fallback:
        return jsonify(fallback)

    return jsonify({"message": "Idle", "running": False})


@app.route("/api/cleanup/deep", methods=["POST"])
def api_deep_clean():
    """Trigger the deep logic-based cleanup"""
    try:
        from tasks import set_status
        set_status("🧹 Deep cleaning database...", True)
        
        def run_clean():
            try:
                conn = get_db()
                cur = conn.cursor()
                cur.execute("SELECT * FROM contacts")
                rows = cur.fetchall()
                
                deleted = 0
                updated = 0
                from processing import ProcessingHandler
                for row in rows:
                    contact = dict(row)
                    contact_id = contact['id']
                    
                    cleaned = ProcessingHandler.process_contact(contact)
                    
                    if cleaned is None:
                        if USE_SQLITE:
                            cur.execute("DELETE FROM contacts WHERE id = ?", (contact_id,))
                        else:
                            cur.execute("DELETE FROM contacts WHERE id = %s", (contact_id,))
                        deleted += 1
                        continue
                        
                    if cleaned.get('phone') != row['phone'] or cleaned.get('email') != row['email']:
                        if USE_SQLITE:
                            cur.execute(
                                "UPDATE contacts SET phone = ?, phone_clean = ?, email = ?, email_valid = ? WHERE id = ?",
                                (cleaned.get('phone'), cleaned.get('phone_clean'), cleaned.get('email'), cleaned.get('email_valid'), contact_id)
                            )
                        else:
                            cur.execute(
                                "UPDATE contacts SET phone = %s, phone_clean = %s, email = %s, email_valid = %s WHERE id = %s",
                                (cleaned.get('phone'), cleaned.get('phone_clean'), cleaned.get('email'), cleaned.get('email_valid'), contact_id)
                            )
                        updated += 1
                
                conn.commit()
                cur.close()
                conn.close()
                set_status("Idle", False)
                return deleted, updated
            except Exception as e:
                set_status("Idle", False)
                raise e

        deleted, updated = run_clean()
        return jsonify({"success": True, "deleted": deleted, "updated": updated})
    except Exception as e:
        logger.error(f"Deep clean failed: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/contact/<int:contact_id>")
def get_contact(contact_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        if USE_SQLITE:
            cur.execute("SELECT * FROM contacts WHERE id = ?", (contact_id,))
        else:
            cur.execute("SELECT * FROM contacts WHERE id = %s", (contact_id,))
        contact = cur.fetchone()
        cur.close()
        conn.close()
        if contact:
            return jsonify(dict(contact))
        return jsonify({"error": "Contact not found"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/logs")
def view_logs():
    try:
        log_files = []
        if LOGS_DIR.exists():
            for f in LOGS_DIR.glob("*.log"):
                log_files.append(
                    {
                        "name": f.name,
                        "size": f.stat().st_size,
                        "modified": f.stat().st_mtime,
                    }
                )
        log_files.sort(key=lambda x: x["modified"], reverse=True)
        return render_template_string(LOGS_HTML, logs=log_files[:30])
    except Exception as e:
        return f"Error reading logs: {e}"


LOG_DETAIL_HTML = """
<!DOCTYPE html>
<html>
<head>
    <title>Log Detail - {{name}}</title>
    <style>
        body { background: #050508; color: #c9d1d9; font-family: 'JetBrains Mono', monospace; margin: 0; padding: 20px; line-height: 1.5; }
        .header { display: flex; justify-content: space-between; align-items: center; border-bottom: 1px solid #2d3148; padding-bottom: 10px; margin-bottom: 20px; }
        .terminal { background: #0a0a0f; padding: 20px; border-radius: 12px; border: 1px solid #2d3148; overflow-x: auto; font-size: 12px; }
        .back-btn { color: #10b981; text-decoration: none; font-size: 14px; font-weight: bold; }
        .back-btn:hover { text-decoration: underline; }
        .line { margin-bottom: 4px; padding-left: 8px; border-left: 2px solid transparent; }
        .ERROR { color: #ef4444; border-left-color: #ef4444; background: rgba(239, 68, 68, 0.05); }
        .SUCCESS { color: #10b981; border-left-color: #10b981; }
        .INFO { color: #94a3b8; }
        .WARNING { color: #f59e0b; border-left-color: #f59e0b; }
        h2 { margin:0; font-size:16px; color: #fff; }
    </style>
</head>
<body>
    <div class="header">
        <a href="/logs" class="back-btn">← Back to Logs</a>
        <h2>{{name}}</h2>
        <div style="font-size: 10px; color: #64748b;">LIVE TAIL (LAST 1000 LINES)</div>
    </div>
    <div class="terminal">
        {% for line in lines %}
        <div class="line {{ 'ERROR' if 'ERROR' in line else 'SUCCESS' if 'SUCCESS' in line else 'WARNING' if 'WARNING' in line else 'INFO' }}">{{ line }}</div>
        {% endfor %}
    </div>
    <script>
        window.scrollTo(0, document.body.scrollHeight);
    </script>
</body>
</html>
"""

@app.route("/logs/<name>")
def get_log(name):
    try:
        log_file = LOGS_DIR / name
        if log_file.exists():
            content = log_file.read_text(errors='replace')
            lines = content.split("\n")
            return render_template_string(LOG_DETAIL_HTML, name=name, lines=lines[-1000:])
        return "Log file not found", 404
    except Exception as e:
        return f"Error reading log detail: {e}", 500


@app.route("/api/trigger/scrape", methods=["POST", "GET"])
def trigger_scrape():
    """Trigger scraping tasks. Supports single (POST JSON) or batch (default)."""
    os.environ.setdefault("CELERY_HEALTH_SERVER_STARTED", "1")
    
    data = {}
    if request.method == "POST":
        try:
            data = request.get_json() or {}
        except:
            data = {}
    
    city = data.get("city") or request.args.get("city")
    category = data.get("category") or request.args.get("category")
    source = data.get("source") or request.args.get("source")
    use_business = data.get("use_business", False)
    auto = bool(data.get("auto")) or request.args.get("auto", "false").lower() == "true"
    
    if not use_business:
        use_business = request.args.get("business", "false").lower() == "true"

    if auto:
        if redis_client:
            redis_client.set("scraper:auto_pilot:active", "1")
        
        set_status(
            "AutoPilot Activated: System will now run continuously...",
            True,
            {"source": "AUTOPILOT", "mode": "CONTINUOUS"},
        )
        task_result = auto_pilot_task.delay()
        if task_result and getattr(task_result, "id", None):
            set_active_task_id(task_result.id)
        msg = "AutoPilot Activated: Continuous scraping cycle started across all cities/categories."
        logger.info("Dashboard activated Continuous AutoPilot")
    elif not city and not category and not source:
        # Fallback for manual button click without params but not 'auto'
        set_status(
            "Queued auto scrape: CA-first official registries...",
            True,
            {"source": "AUTO", "category": "Chartered Accountants"},
        )
        task_result = direct_gov_scrape_batch.delay()
        if task_result and getattr(task_result, "id", None):
            set_active_task_id(task_result.id)
        msg = "Auto scrape queued: Chartered Accountants first, then official registries."
        logger.info("Dashboard triggered CA-first auto government scrape")
    elif city and category:
        log_msg = f"Dashboard triggered manual scrape: {category} in {city} (Source: {source or 'Auto'})"
        set_status(
            f"Queued: {category} in {city}...",
            True,
            {"city": city, "category": category, "source": source or "QUEUE"},
        )
        task_result = scrape_category_task.delay(city=city, category=category, source=source, use_business=use_business)
        if task_result and getattr(task_result, "id", None):
            set_active_task_id(task_result.id)
        msg = f"🚀 Scrape queued for {category} in {city}!"
        logger.info(log_msg)
    else:
        set_status(
            "Queued batch fast-scrape for all configured targets...",
            True,
            {"source": source or "QUEUE"},
        )
        task_result = fast_scrape_task.delay(source=source)
        if task_result and getattr(task_result, "id", None):
            set_active_task_id(task_result.id)
        msg = f"🚀 Batch fast-scrape queued for all Official sources!"
    
    return jsonify({"message": msg, "task_id": getattr(task_result, "id", None)})

@app.route("/api/trigger/stop", methods=["POST", "GET"])
def stop_scrape():
    """Stop active scraping tasks and AutoPilot"""
    # 1. Stop AutoPilot flag
    if redis_client:
        try:
            redis_client.set("scraper:auto_pilot:active", "0")
        except Exception:
            pass
            
    # 2. Find and revoke active celery task
    task_id = get_active_task_id()
    revoked = False
    if task_id:
        try:
            from tasks import celery_app
            celery_app.control.revoke(task_id, terminate=True, signal='SIGTERM')
            revoked = True
            logger.info(f"Successfully revoked celery task {task_id}")
        except Exception as e:
            logger.warning(f"Could not revoke celery task {task_id}: {e}")
            
    # 3. Clear active task ID
    set_active_task_id(None)
    
    # 4. Reset status
    msg = "Scraper STOP signal sent. "
    if revoked:
        msg += f"Active task {task_id} terminated successfully."
    else:
        msg += "No active celery task was found or running."
        
    set_status(msg, False)
    return jsonify({"success": True, "message": msg})


@app.route("/api/trigger/fast-scrape", methods=["POST"])
def trigger_fast_scrape():
    """Trigger fast parallel scraping with higher concurrency"""
    os.environ.setdefault("CELERY_HEALTH_SERVER_STARTED", "1")
    try:
        max_concurrent = request.args.get("concurrency", 5, type=int)
        set_status(
            f"Queued fast scrape with concurrency={max_concurrent}...",
            True,
            {"source": "QUEUE", "concurrency": max_concurrent},
        )
        task_result = fast_scrape_task.delay(max_concurrent=max_concurrent)
        if task_result and getattr(task_result, "id", None):
            set_active_task_id(task_result.id)
        return jsonify({
            "message": f"⚡ Fast scrape queued with concurrency={max_concurrent}!",
            "task_id": getattr(task_result, "id", None),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/trigger/direct-scrape", methods=["POST", "GET"])
def trigger_direct_scrape():
    """
    Trigger direct scraping WITHOUT proxy - for government sites.
    Uses polite HTTP fetching to avoid blocking.
    """
    os.environ.setdefault("CELERY_HEALTH_SERVER_STARTED", "1")
    try:
        source = request.args.get("source", "SEBI") or request.json.get("source", "SEBI") if request.is_json else "SEBI"
        city = request.args.get("city") or (request.json.get("city") if request.is_json else None)
        category = request.args.get("category") or (request.json.get("category") if request.is_json else None)
        
        set_status(
            f"Queued direct scrape: {source}...",
            True,
            {"source": source, "city": city, "category": category},
        )

        # School scraping needs more time - use dedicated task
        if source.upper() == "SCHOOL":
            from tasks import school_scrape_task
            task_result = school_scrape_task.delay(source=source, city=city, category=category)
        else:
            task_result = direct_scrape_task.delay(source=source, city=city, category=category)
        if task_result and getattr(task_result, "id", None):
            set_active_task_id(task_result.id)
        
        return jsonify({
            "message": f"🔓 Direct scrape queued for {source}!",
            "task_id": getattr(task_result, "id", None),
            "source": source,
            "city": city,
            "category": category,
        })
    except Exception as e:
        logger.error(f"Direct scrape trigger error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/trigger/direct-gov-batch", methods=["POST", "GET"])
def trigger_direct_gov_batch():
    """
    Trigger batch direct scraping for all government sites.
    Scrapes SEBI, ICAI, NSE, MCA, AMFI without proxies.
    """
    os.environ.setdefault("CELERY_HEALTH_SERVER_STARTED", "1")
    try:
        set_status(
            "Queued direct government sites batch...",
            True,
            {"source": "GOVERNMENT"},
        )
        
        task_result = direct_gov_scrape_batch.delay()
        if task_result and getattr(task_result, "id", None):
            set_active_task_id(task_result.id)
        
        return jsonify({
            "message": "🔓 Direct gov batch queued! Scraping ICAI (CAs), AMFI (MF Agents), SEBI (Investment Advisors), IRDAI (Insurance Agents)...",
            "task_id": getattr(task_result, "id", None),
        })
    except Exception as e:
        logger.error(f"Direct gov batch trigger error: {e}")
        return jsonify({"error": str(e)}), 500



@app.route("/api/contacts")
def api_contacts():
    try:
        conn = get_db()
        cur = conn.cursor()

        page = request.args.get("page", 1, type=int)
        limit = min(request.args.get("limit", 100, type=int), 1000)
        offset = (page - 1) * limit

        search_query = request.args.get("q", "")
        filter_city = request.args.get("city", "")
        filter_category = request.args.get("category", "")
        filter_source = request.args.get("source", "")
        filter_quality = request.args.get("quality", "")
        sort_by = request.args.get("sort", "date")

        where_sql, params = build_contact_filters(
            search_query, filter_city, filter_category, filter_source
        )

        if filter_quality:
            try:
                q_min = int(filter_quality)
                where_sql += f" AND quality_score >= {db_placeholder()}"
                params.append(q_min)
            except ValueError:
                pass

        sort_map = {
            "name": "LOWER(name) ASC",
            "name_desc": "LOWER(name) DESC",
            "date": "scraped_at DESC",
            "date_asc": "scraped_at ASC",
            "score": "quality_score DESC",
            "score_asc": "quality_score ASC",
            "city": "LOWER(city) ASC",
            "city_desc": "LOWER(city) DESC",
            "category": "LOWER(category) ASC",
            "category_desc": "LOWER(category) DESC",
            "source": "LOWER(source) ASC",
            "source_desc": "LOWER(source) DESC",
            "phone": "phone ASC",
            "phone_desc": "phone DESC",
            "email": "email ASC",
            "email_desc": "email DESC",
        }
        order_clause = sort_map.get(sort_by, "scraped_at DESC")

        ph = db_placeholder()
        query = f"SELECT id, name, phone, email, address, city, area, state, category, source, source_url, quality_score, quality_tier, scraped_at FROM contacts WHERE {where_sql} ORDER BY {order_clause} LIMIT {ph} OFFSET {ph}"
        count_query = f"SELECT COUNT(*) as cnt FROM contacts WHERE {where_sql}"

        cur.execute(query, params + [limit, offset])
        contacts = cur.fetchall()
        cur.execute(count_query, params)
        total = cur.fetchone()["cnt"]

        cur.execute(f"SELECT COUNT(*) as cnt FROM contacts WHERE {where_sql} AND phone IS NOT NULL AND phone != ''", params)
        phone_count = cur.fetchone()["cnt"]
        cur.execute(f"SELECT COUNT(*) as cnt FROM contacts WHERE {where_sql} AND email IS NOT NULL AND email != ''", params)
        email_count = cur.fetchone()["cnt"]

        cur.close()
        conn.close()
        return jsonify(
            {
                "total": total,
                "page": page,
                "limit": limit,
                "total_pages": max(1, (total + limit - 1) // limit) if total else 1,
                "filtered_total": total,
                "contacts": [dict(c) for c in contacts],
                "stats": {"total": total, "phone": phone_count, "email": email_count},
            }
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/merge-local", methods=["POST"])
def api_merge_local():
    """Accept locally-exported contacts and merge into the active database."""
    try:
        data = request.get_json(force=True)
        contacts = data.get("contacts", [])
        if not contacts:
            return jsonify({"error": "No contacts provided"}), 400

        conn = get_db()
        cur = conn.cursor()
        saved = 0
        skipped = 0

        # Determine if we're on SQLite or PostgreSQL
        is_sqlite = USE_SQLITE or isinstance(conn, sqlite3.Connection)

        if is_sqlite:
            for c in contacts:
                try:
                    cur.execute(
                        """INSERT OR IGNORE INTO contacts
                        (name, phone, email, address, category, city, area, state, source, source_url,
                         phone_clean, email_valid, enriched, arn, license_no, membership_no,
                         quality_score, quality_tier, scraped_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (
                            c.get("name", ""), c.get("phone", ""), c.get("email", ""),
                            c.get("address", ""), c.get("category", ""), c.get("city", ""),
                            c.get("area", ""), c.get("state", ""), c.get("source", ""),
                            c.get("source_url", ""), c.get("phone_clean", ""),
                            1 if c.get("email_valid") else 0, 1 if c.get("enriched") else 0,
                            c.get("arn", ""), c.get("license_no", ""), c.get("membership_no", ""),
                            c.get("quality_score", 0), c.get("quality_tier", "low"),
                            c.get("scraped_at"),
                        ),
                    )
                    if cur.rowcount > 0:
                        saved += 1
                    else:
                        skipped += 1
                except Exception:
                    skipped += 1
        else:
            from psycopg2.extras import execute_values
            rows = []
            for c in contacts:
                rows.append((
                    c.get("name", ""), c.get("phone", ""), c.get("email", ""),
                    c.get("address", ""), c.get("category", ""), c.get("city", ""),
                    c.get("area", ""), c.get("state", ""), c.get("source", ""),
                    c.get("source_url", ""), c.get("phone_clean", ""),
                    c.get("arn", ""), c.get("license_no", ""), c.get("membership_no", ""),
                    c.get("quality_score", 0), c.get("quality_tier", "low"),
                    c.get("scraped_at"),
                ))
            if rows:
                execute_values(cur,
                    """INSERT INTO contacts
                    (name, phone, email, address, category, city, area, state, source, source_url,
                     phone_clean, arn, license_no, membership_no, quality_score, quality_tier, scraped_at)
                    VALUES %s
                    ON CONFLICT (phone_clean) WHERE phone_clean IS NOT NULL DO NOTHING""",
                    rows,
                )
                saved = cur.rowcount if cur.rowcount > 0 else len(rows)

        conn.commit() if not conn.autocommit else None
        cur.close()
        conn.close()
        return jsonify({"saved": saved, "total": len(contacts), "db": "sqlite" if is_sqlite else "postgresql"})
    except Exception as e:
        logger.error(f"Merge local failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/export/<fmt>")
def export(fmt):
    logger.info(f"Export requested: format={fmt}, args={dict(request.args)}")
    
    if fmt not in ("csv", "excel", "json", "pdf", "xlsx"):
        return "Invalid format. Use csv, excel, xlsx, json, or pdf.", 400
    
    export_all = request.args.get("all") == "true"
    target_cat = request.args.get("category", "")
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        schools_only = request.args.get("schools_only") == "true"
        financial_only = request.args.get("financial_only") == "true"

        if export_all:
            search_query = request.args.get("q", "")
            filter_city = request.args.get("city", "")
            filter_source = request.args.get("source", "")
            filter_quality = request.args.get("quality", "")
            filter_category = target_cat or request.args.get("category", "")
            where_sql, params = build_contact_filters(
                search_query, filter_city, filter_category, filter_source,
                quality=filter_quality,
                exclude_schools=financial_only, only_schools=schools_only,
            )
        else:
            search_query = request.args.get("q", "")
            filter_city = request.args.get("city", "")
            filter_category = request.args.get("category", "")
            filter_source = request.args.get("source", "")
            where_sql, params = build_contact_filters(
                search_query,
                filter_city,
                filter_category,
                filter_source,
                exclude_schools=financial_only,
                only_schools=schools_only
            )
            
        logger.info(f"Export query: WHERE {where_sql} with params {params}")

        # For large exports, we use streaming to avoid memory issues
        if fmt == "csv":
            import csv
            
            def generate():
                # Yield Header
                out = io.StringIO()
                fields = ["name", "phone", "email", "address", "category", "city", "area", "state", "source", "scraped_at", "arn", "license_no"]
                writer = csv.DictWriter(out, fieldnames=fields, extrasaction='ignore')
                writer.writeheader()
                yield out.getvalue()
                
                # Yield Data in chunks
                cur.execute(f"SELECT * FROM contacts WHERE {where_sql} ORDER BY scraped_at DESC", params)
                while True:
                    rows = cur.fetchmany(1000)
                    if not rows:
                        break
                    
                    out = io.StringIO()
                    writer = csv.DictWriter(out, fieldnames=fields, extrasaction='ignore')
                    for r in rows:
                        row = dict(r)
                        for k, v in row.items():
                            if isinstance(v, (datetime, date)):
                                row[k] = v.isoformat()
                            elif v is None:
                                row[k] = ""
                        writer.writerow(row)
                    yield out.getvalue()
                
                cur.close()
                conn.close()

            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            if export_all and target_cat:
                filename = f"bulk_export_{target_cat.replace(' ', '_').lower()}_{ts}.csv"
            else:
                filename = f"export_{ts}.csv" if not export_all else f"bulk_export_{ts}.csv"
            
            return Response(
                stream_with_context(generate()),
                mimetype="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )

        # For non-streaming formats, we still need to load data
        # We cap these to prevent crash
        limit = 50000 if export_all else 20000
        cur.execute(f"SELECT * FROM contacts WHERE {where_sql} ORDER BY scraped_at DESC LIMIT {limit}", params)
        rows = [dict(r) for r in cur.fetchall()]
        cur.close()
        conn.close()
        logger.info(f"Export found {len(rows)} rows")
        
    except Exception as e:
        logger.error(f"Export error: {e}")
        return jsonify({"error": str(e)}), 500

    if not rows:
        rows = []

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    parts = [ts]
    if not export_all:
        if filter_city: parts.append(filter_city.replace(' ', '_'))
        if filter_category: parts.append(filter_category.replace(' ', '_'))
        if search_query: parts.append(search_query.replace(' ', '_'))
    else:
        if target_cat:
            parts.append(f"BULK_{target_cat.replace(' ', '_').upper()}")
        else:
            parts.append("BULK_ALL")
        
    filename_prefix = "_".join(parts)
    total_rows = len(rows)

    if fmt == "json":
        return jsonify({
            "status": "success",
            "count": total_rows,
            "filters": "ALL" if export_all else {
                "search": search_query,
                "city": filter_city,
                "category": filter_category,
                "source": filter_source
            },
            "timestamp": datetime.now().isoformat(),
            "data": rows
        })

    if fmt in ("excel", "xlsx"):
        logger.info(f"Starting Excel generation for {len(rows)} records...")
        try:
            wb = Workbook(write_only=True)
            ws = wb.create_sheet("Intelligence Data")
            
            if rows:
                headers = list(rows[0].keys())
                ws.append(headers)
                
                for r in rows:
                    clean_row = []
                    for k in headers:
                        v = r.get(k, "")
                        if isinstance(v, (datetime, date)):
                            clean_row.append(v.strftime('%Y-%m-%d %H:%M:%S') if isinstance(v, datetime) else v.strftime('%Y-%m-%d'))
                        elif v is None:
                            clean_row.append("")
                        else:
                            if not isinstance(v, (int, float, str, bool)):
                                clean_row.append(str(v))
                            else:
                                clean_row.append(v)
                    ws.append(clean_row)
            
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            
            return send_file(
                out, 
                download_name=f"{filename_prefix}_{total_rows}rows.xlsx", 
                as_attachment=True, 
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as excel_err:
            logger.error(f"CRITICAL EXCEL ERROR: {excel_err}")
            return f"Excel Export Failed: {str(excel_err)}", 500

    if fmt == "pdf":
        # Keep PDF logic as is, limited to 1000 rows
        from fpdf import FPDF
        def clean_pdf_text(text):
            if not text: return ""
            try: return str(text).encode('latin-1', 'replace').decode('latin-1')
            except Exception: return ""

        class PDF(FPDF):
            def header(self):
                self.set_font('helvetica', 'B', 15)
                self.cell(0, 10, 'Maysan Labs Intelligence Export', border=False, align='C')
                self.ln(10)
            def footer(self):
                self.set_y(-15)
                self.set_font('helvetica', 'I', 8)
                self.cell(0, 10, f'Page {self.page_no()}', align='C')

        pdf = PDF(orientation='L', unit='mm', format='A4')
        pdf.add_page()
        pdf.set_font('helvetica', 'B', 9)
        cols = ["Name", "Phone", "Email", "City", "Category", "Source"]
        col_widths = [60, 40, 60, 35, 45, 35]
        for i, col in enumerate(cols): pdf.cell(col_widths[i], 7, col, 1)
        pdf.ln()
        pdf.set_font('helvetica', '', 8)
        for r in rows[:1000]:
            pdf.cell(col_widths[0], 6, clean_pdf_text(r.get("name"))[:35], 1)
            pdf.cell(col_widths[1], 6, clean_pdf_text(r.get("phone"))[:20], 1)
            pdf.cell(col_widths[2], 6, clean_pdf_text(r.get("email"))[:35], 1)
            pdf.cell(col_widths[3], 6, clean_pdf_text(r.get("city"))[:20], 1)
            pdf.cell(col_widths[4], 6, clean_pdf_text(r.get("category"))[:25], 1)
            pdf.cell(col_widths[5], 6, clean_pdf_text(r.get("source"))[:20], 1)
            pdf.ln()
            
        pdf_data = pdf.output()
        if isinstance(pdf_data, str): pdf_data = pdf_data.encode('latin-1', 'replace')
        return send_file(io.BytesIO(pdf_data), download_name=f"{filename_prefix}_sample.pdf", as_attachment=True, mimetype="application/pdf")

    return "Invalid format", 400


LOGS_HTML = """
<!DOCTYPE html>
<html>
<head>
    <title>Scraper Logs</title>
    <style>
        body { background: #0d1117; color: #c9d1d9; font-family: monospace; padding: 20px; }
        h1 { color: #fff; }
        .log-list { list-style: none; padding: 0; }
        .log-list li { padding: 10px; border-bottom: 1px solid #2d3148; }
        .log-list a { color: #58a6ff; text-decoration: none; }
        .log-list a:hover { text-decoration: underline; }
        .log-content { background: #161824; padding: 20px; border-radius: 8px; overflow-x: auto; white-space: pre-wrap; font-size: 12px; max-height: 70vh; }
        .back { color: #8b8fa3; margin-bottom: 20px; }
    </style>
</head>
<body>
    <h1>Scraper Logs</h1>
    <a class="back" href="/">← Back to Dashboard</a>
    {% if logs %}
    <ul class="log-list">
    {% for log in logs %}
        <li><a href="/logs/{{log.name}}">{{log.name}}</a> - {{(log.size/1024)|round(1)}} KB</li>
    {% endfor %}
    </ul>
    {% else %}
    <p>No logs found.</p>
    {% endif %}
</body>
</html>
"""


@app.route("/api/cleanup/empty", methods=["DELETE"])
def cleanup_empty_contacts():
    """Delete contacts that have neither phone nor email"""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            DELETE FROM contacts 
            WHERE (phone IS NULL OR TRIM(phone) = '') 
            AND (email IS NULL OR TRIM(email) = '')
        """)
        deleted_count = cur.rowcount
        conn.commit()
        cur.execute("SELECT COUNT(*) as cnt FROM contacts")
        remaining = cur.fetchone()["cnt"]
        cur.close()
        conn.close()
        return jsonify({
            "success": True,
            "deleted": deleted_count,
            "remaining": remaining,
            "message": f"Deleted {deleted_count} contacts with no phone or email",
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/cleanup/quality", methods=["POST"])
def cleanup_low_quality():
    """Recalculate and update quality scores for all contacts"""
    try:
        from processing import ProcessingHandler
        from tasks import set_status
        set_status("🔍 Auditing lead quality...", True)
        
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM contacts")
        contacts = cur.fetchall()
        if not contacts:
            set_status("Idle", False)
            return jsonify({"success": True, "updated": 0, "message": "No contacts to update"})
            
        updated = 0
        for contact in contacts:
            try:
                processed = ProcessingHandler.process_contact(dict(contact))
                if USE_SQLITE:
                    cur.execute("""
                        UPDATE contacts 
                        SET phone_clean = ?, 
                            email_valid = ?, 
                            quality_score = ?, 
                            quality_tier = ?
                        WHERE id = ?
                    """, (processed.get("phone_clean"), processed.get("email_valid", False), processed.get("quality_score", 0), processed.get("quality_tier", "low"), contact["id"]))
                else:
                    cur.execute("""
                        UPDATE contacts 
                        SET phone_clean = %s, 
                            email_valid = %s, 
                            quality_score = %s, 
                            quality_tier = %s
                        WHERE id = %s
                    """, (processed.get("phone_clean"), processed.get("email_valid", False), processed.get("quality_score", 0), processed.get("quality_tier", "low"), contact["id"]))
                updated += 1
                
                # Periodically commit and update status for very large sets
                if updated % 500 == 0:
                    conn.commit()
                    set_status(f"🔍 Audited {updated} leads...", True)
                    
            except Exception:
                continue
                
        conn.commit()
        cur.close()
        conn.close()
        set_status("Idle", False)
        return jsonify({"success": True, "updated": updated})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
@app.route("/api/maintenance/normalize", methods=["POST"])
def api_maintenance_normalize():
    """Trigger system-wide category normalization"""
    try:
        from processing import ProcessingHandler
        from tasks import set_status
        set_status("🧹 Normalizing all categories...", True)
        
        conn = get_db()
        stats = ProcessingHandler.clean_database_logic(conn)
        conn.close()
        
        set_status("Idle", False)
        return jsonify({"success": True, **stats})
    except Exception as e:
        logger.error(f"Maintenance failed: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/stream/stats")
def stream_stats():
    """Server-Sent Events endpoint for live stats updates"""
    schools_only = request.args.get("schools_only") == "true"
    search_q = request.args.get("q", "")
    filter_city = request.args.get("city", "")
    filter_cat = request.args.get("category", "")
    filter_src = request.args.get("source", "")
    filter_qual = request.args.get("quality", "")
    def generate():
        while True:
            try:
                conn = get_db()
                cur = conn.cursor()
                where_filter, filter_params = build_contact_filters(
                    search_q, filter_city, filter_cat, filter_src,
                    quality=filter_qual,
                    exclude_schools=not schools_only,
                    only_schools=schools_only,
                )

                cur.execute(f"SELECT COUNT(*) as cnt FROM contacts WHERE {where_filter}", filter_params)
                total = cur.fetchone()["cnt"]
                
                cur.execute(f"SELECT COUNT(*) as cnt FROM contacts WHERE {where_filter} AND phone_clean IS NOT NULL AND phone_clean <> ''", filter_params)
                with_phone = cur.fetchone()["cnt"]
                
                cur.execute(f"SELECT COUNT(*) as cnt FROM contacts WHERE {where_filter} AND email IS NOT NULL AND email <> ''", filter_params)
                with_email = cur.fetchone()["cnt"]

                status_data = {}
                if redis_client:
                    try:
                        raw_status = redis_client.get("scraper_status")
                        if raw_status: status_data = json.loads(raw_status)
                    except: pass
                if not status_data:
                    try:
                        cur.execute("SELECT value FROM system_status WHERE key = 'scraper_status'")
                        row = cur.fetchone()
                        if row: status_data = json.loads(row["value"])
                    except: pass
                cur.execute("SELECT * FROM scraper_logs ORDER BY created_at DESC LIMIT 15")
                logs = [dict(r) for r in cur.fetchall()]
                for l in logs:
                    if isinstance(l['created_at'], datetime):
                        l['time'] = l['created_at'].strftime("%H:%M:%S")
                    else:
                        l['time'] = str(l['created_at'])[-8:]
                yield f"data: {json.dumps({'total': total, 'with_phone': with_phone, 'with_email': with_email, 'scraper_status': status_data, 'activity_logs': logs}, cls=CustomJSONEncoder)}\n\n"
                cur.close()
                conn.close()
            except Exception as e:
                logger.error(f"Stream error: {e}")
            time.sleep(2)
    return Response(generate(), mimetype="text/event-stream")


@app.route("/api/stats/charts")
def api_chart_stats():
    try:
        schools_only = request.args.get("schools_only") == "true"
        ph = "?" if USE_SQLITE else "%s"
        
        if schools_only:
            where_sql = f"WHERE (LOWER(category) LIKE {ph} OR LOWER(source) IN ('npsc', 'bsai', 'aisa'))"
            params = ["%school%"]
        else:
            where_sql = f"WHERE (LOWER(category) NOT LIKE {ph} AND LOWER(source) NOT IN ('npsc', 'bsai', 'aisa'))"
            params = ["%school%"]
            
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute(f"SELECT source, COUNT(*) as count FROM contacts {where_sql} GROUP BY source", params)
        sources = [dict(r) for r in cur.fetchall()]
        
        # Normalize categories in Python to ensure perfect grouping regardless of DB state
        cur.execute(f"SELECT category, COUNT(*) as count FROM contacts {where_sql} GROUP BY category", params)
        raw_cats = cur.fetchall()
        from processing import ProcessingHandler
        cat_map = {}
        for r in raw_cats:
            norm = ProcessingHandler.normalize_category(r["category"])
            cat_map[norm] = cat_map.get(norm, 0) + r["count"]
        
        categories = [{"category": k, "count": v} for k, v in sorted(cat_map.items(), key=lambda x: x[1], reverse=True)[:10]]
 
        if USE_SQLITE:
            cur.execute(f"SELECT strftime('%Y-%m-%d', scraped_at) as date, COUNT(*) as count FROM contacts {where_sql} GROUP BY date ORDER BY date DESC LIMIT 7", params)
        else:
            cur.execute(f"SELECT scraped_at::date as date, COUNT(*) as count FROM contacts {where_sql} GROUP BY date ORDER BY date DESC LIMIT 7", params)
        trend = [dict(r) for r in cur.fetchall()]
        cur.close()
        conn.close()
        return jsonify({"sources": sources, "categories": categories, "trend": trend[::-1]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(PORT), debug=True)
